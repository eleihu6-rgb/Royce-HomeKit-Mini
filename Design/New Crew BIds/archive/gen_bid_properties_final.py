#!/usr/bin/env python3
"""
Read bid_properties-definition-Final.xlsx, reassign IDs with the scheme:
  Pairing  → 101, 102, 103 …
  DaysOff  → 201, 202, 203 …
  Reserve  → 301, 302, 303 …
  Line     → 401, 402, 403 …
Row order preserved. Save as timestamped file.
"""
from pathlib import Path
import openpyxl, datetime, copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path("/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds")
SRC  = BASE / "bid_properties-definition-Final.xlsx"

TYPE_BASE = {'Pairing': 100, 'DaysOff': 200, 'Reserve': 300, 'Line': 400}

# ── read source ───────────────────────────────────────────────────────────────
wb_src = openpyxl.load_workbook(SRC)
ws_src = wb_src.active
h = {ws_src.cell(1, c).value: c for c in range(1, ws_src.max_column + 1)}
cols = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]
print(f"Source: {SRC.name}")
print(f"Columns: {cols}")

# Read header, guide, and data rows
guide_row = [ws_src.cell(2, c).value for c in range(1, ws_src.max_column + 1)]
data_rows = []
for r in range(3, ws_src.max_row + 1):
    row = [ws_src.cell(r, c).value for c in range(1, ws_src.max_column + 1)]
    data_rows.append(row)

# ── reassign IDs ──────────────────────────────────────────────────────────────
id_col_idx  = cols.index('id')
type_col_idx = cols.index('bid_type')

type_counters = {'Pairing': 0, 'DaysOff': 0, 'Reserve': 0, 'Line': 0}
for row in data_rows:
    btype = row[type_col_idx]
    if btype in type_counters:
        type_counters[btype] += 1
        row[id_col_idx] = TYPE_BASE[btype] + type_counters[btype]

print(f"\nID counts: { {t: type_counters[t] for t in TYPE_BASE} }")

# ── styles ────────────────────────────────────────────────────────────────────
thin   = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

GUIDE_FONT  = Font(italic=True, name='Calibri', size=10, color='595959')
GUIDE_FILL  = PatternFill('solid', fgColor='FFF2CC')
GUIDE_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

TYPE_FILL = {
    'Pairing': PatternFill('solid', fgColor='E2EFDA'),  # light green
    'DaysOff': PatternFill('solid', fgColor='FCE4D6'),  # light orange
    'Reserve': PatternFill('solid', fgColor='FFF2CC'),  # light yellow
    'Line':    PatternFill('solid', fgColor='EDEDED'),  # light grey
}
DATA_FONT  = Font(name='Calibri', size=10)
DATA_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
ID_ALIGN   = Alignment(horizontal='center', vertical='top')

COL_WIDTHS = {
    'id':              7,
    'bid_type':       10,
    'N-PBS Property': 38,
    'display_name':   40,
    'any_or_every':   20,
    'operator':       28,
    'validation_json':55,
    'notes':          42,
    'crew_count':      9,
}

# ── build workbook ────────────────────────────────────────────────────────────
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = 'bid_properties'

# Row 1 — header
for c, col_name in enumerate(cols, 1):
    cell = ws.cell(1, c, value=col_name)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = HDR_ALIGN
    cell.border    = border
ws.row_dimensions[1].height = 22

# Row 2 — guide
for c, val in enumerate(guide_row, 1):
    cell = ws.cell(2, c, value=val)
    cell.font      = GUIDE_FONT
    cell.fill      = GUIDE_FILL
    cell.alignment = GUIDE_ALIGN
    cell.border    = border
ws.row_dimensions[2].height = 40

# Rows 3+ — data
for r_idx, row in enumerate(data_rows, 3):
    btype = row[type_col_idx]
    fill  = TYPE_FILL.get(btype, PatternFill())
    for c, val in enumerate(row, 1):
        cell = ws.cell(r_idx, c, value=val)
        cell.font   = DATA_FONT
        cell.fill   = fill
        cell.border = border
        if c == id_col_idx + 1:
            cell.alignment = ID_ALIGN
        else:
            cell.alignment = DATA_ALIGN

# Column widths
for c, col_name in enumerate(cols, 1):
    width = COL_WIDTHS.get(col_name, 14)
    ws.column_dimensions[get_column_letter(c)].width = width

ws.freeze_panes = 'A3'

# ── save ──────────────────────────────────────────────────────────────────────
ts  = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f"bid_properties-definition-{ts}.xlsx"
wb.save(out)
print(f"Saved: {out.name}")

# ── verify IDs ────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active
print("\nID assignments:")
prev_type = None
for r in range(3, ws2.max_row + 1):
    pid   = ws2.cell(r, id_col_idx + 1).value
    btype = ws2.cell(r, type_col_idx + 1).value
    name  = ws2.cell(r, cols.index('display_name') + 1).value
    if btype != prev_type:
        print(f"\n  [{btype}]")
        prev_type = btype
    print(f"    {pid}  {name}")
