#!/usr/bin/env python3
"""
Generate cleaned final Excel:
  1. Remove rows never seen in 4-month data: 131, 132, 133 (Pairing), 207 (DaysOff)
  2. Fix operator column — remove false "In" entries caused by qualifier artifacts
     (Counting Deadhead Legs / PairingStage Beginning)
  3. Rebuild with full styling preserved
"""
import json, datetime, copy
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path('/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds')
SRC  = BASE / 'bid_properties-definition-2026-03-16-060755.xlsx'

# ── rows to drop (new_id) ─────────────────────────────────────────────────────
DROP_IDS = {131, 132, 133, 207}

# ── corrected operator values keyed by new_id ─────────────────────────────────
# Based on 4-month (May/Jul/Aug/Dec) actual bid analysis.
# "In" removed from 107, 108, 122, 124 — those were qualifier artifacts.
OPERATOR = {
    # Pairing
    101: ["In"],
    102: ["In"],
    103: ["<", "=", ">", "Between"],
    104: ["In"],
    105: ["<", "=", ">", "Between"],
    106: ["Between", "In"],
    107: ["<", "=", ">"],          # was ["<","=",">","In"] — "In" was artifact
    108: ["<", "=", ">"],          # was ["=",">","<","In"] — "In" was artifact
    109: ["<", "=", ">", "Between"],
    110: ["Between", "In"],
    111: ["<", "=", ">", "Between"],
    112: ["<", "=", ">", "Between"],
    113: ["<", ">", "Between"],
    114: ["<", "=", ">", "Between"],
    115: ["In"],
    116: ["In"],
    117: None,                     # boolean flag — no operator
    118: ["<", ">", "Between"],
    119: ["<", ">", "Between"],
    120: ["<", "=", ">", "Between"],
    121: ["<", ">"],
    122: ["<", "=", ">", "Between"],  # was also "In" — artifact from PairingStage
    123: ["Between", "In"],
    124: ["<", ">"],               # was ["In",">","<"] — "In" was artifact
    125: ["<", ">"],
    126: ["<", ">", "Between"],
    127: ["=", ">", "Between"],
    128: None,                     # boolean flag — no operator
    129: ["<", ">", "Between"],
    130: [">"],
    # DaysOff
    201: ["In", "Between"],
    202: None,
    203: None,
    204: ["In", "Between"],
    205: ["Between"],
    206: ["In"],
    # Reserve
    301: ["In"],
    302: ["In"],
    # Line
    401: None,
    402: None,
    403: None,
    404: None,
    405: None,
    406: ["In"],
    407: ["In"],
}

# ── load source ───────────────────────────────────────────────────────────────
print(f'Source: {SRC.name}')
wb_src = openpyxl.load_workbook(SRC)
ws_src = wb_src.active
h = {ws_src.cell(1, c).value: c for c in range(1, ws_src.max_column + 1)}
cols = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]
id_col   = h['id']
op_col   = h['operator']
bt_col   = h['bid_type']

# ── collect rows to keep ──────────────────────────────────────────────────────
guide_row  = [ws_src.cell(2, c).value for c in range(1, ws_src.max_column + 1)]
guide_styl = [(ws_src.cell(2, c).font, ws_src.cell(2, c).fill,
               ws_src.cell(2, c).alignment, ws_src.cell(2, c).border)
              for c in range(1, ws_src.max_column + 1)]

data_rows  = []   # (values_list, styles_list, new_id)
for r in range(3, ws_src.max_row + 1):
    new_id = ws_src.cell(r, id_col).value
    if new_id in DROP_IDS:
        continue
    vals   = [ws_src.cell(r, c).value for c in range(1, ws_src.max_column + 1)]
    styles = [(ws_src.cell(r, c).font, ws_src.cell(r, c).fill,
               ws_src.cell(r, c).alignment, ws_src.cell(r, c).border)
              for c in range(1, ws_src.max_column + 1)]
    # Apply corrected operator
    if new_id in OPERATOR:
        op_val = OPERATOR[new_id]
        vals[op_col - 1] = json.dumps(op_val) if op_val is not None else None
    data_rows.append((vals, styles, new_id))

# ── styles ────────────────────────────────────────────────────────────────────
thin   = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

TYPE_FILL = {
    'Pairing': PatternFill('solid', fgColor='E2EFDA'),
    'DaysOff': PatternFill('solid', fgColor='FCE4D6'),
    'Reserve': PatternFill('solid', fgColor='FFF2CC'),
    'Line':    PatternFill('solid', fgColor='EDEDED'),
}

COL_WIDTHS = {
    'id':              7,
    'bid_type':       10,
    'N-PBS Property': 38,
    'display_name':   40,
    'award_or_avoid': 18,
    'any_or_every':   18,
    'operator':       28,
    'validation_json':58,
    'notes':          42,
    'crew_count':      9,
}

DATA_FONT  = Font(name='Calibri', size=10)
DATA_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
ID_ALIGN   = Alignment(horizontal='center', vertical='top')

GUIDE_FONT  = Font(italic=True, name='Calibri', size=10, color='595959')
GUIDE_FILL  = PatternFill('solid', fgColor='FFF2CC')
GUIDE_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ── build new workbook ────────────────────────────────────────────────────────
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = 'bid_properties'

# Row 1 — header
for c, col_name in enumerate(cols, 1):
    cell = ws.cell(1, c, value=col_name)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = HDR_ALIGN
    cell.border    = border
ws.row_dimensions[1].height = 22

# Row 2 — guide
for c, val in enumerate(guide_row, 1):
    cell = ws.cell(2, c, value=val)
    cell.font      = GUIDE_FONT
    cell.fill      = GUIDE_FILL
    cell.alignment = GUIDE_ALIGN
    cell.border    = border
ws.row_dimensions[2].height = 40

# Data rows
bt_idx = cols.index('bid_type')
id_idx = cols.index('id')

for r_idx, (vals, styles, new_id) in enumerate(data_rows, 3):
    btype = vals[bt_idx]
    fill  = TYPE_FILL.get(btype, PatternFill())
    for c, val in enumerate(vals, 1):
        cell = ws.cell(r_idx, c, value=val)
        cell.font   = DATA_FONT
        cell.fill   = fill
        cell.border = border
        if c == id_idx + 1:
            cell.alignment = ID_ALIGN
        else:
            cell.alignment = DATA_ALIGN

# Column widths
for c, col_name in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS.get(col_name, 14)

ws.freeze_panes = 'A3'

# ── save ──────────────────────────────────────────────────────────────────────
ts  = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f'bid_properties-definition-{ts}.xlsx'
wb.save(out)
print(f'Saved: {out.name}')
print(f'Rows: {len(data_rows)} data rows  ({len(DROP_IDS)} dropped)')

# ── verify ────────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active
h2  = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}

from collections import Counter
type_counts = Counter()
for r in range(3, ws2.max_row + 1):
    type_counts[ws2.cell(r, h2['bid_type']).value] += 1

print(f'\nRow counts by type: {dict(type_counts)}')
print(f'\nColumns: {list(h2.keys())}')

print('\nOperator spot-check (fixed rows):')
check_ids = [107, 108, 122, 124]
for r in range(3, ws2.max_row + 1):
    nid = ws2.cell(r, h2['id']).value
    if nid in check_ids:
        dn = ws2.cell(r, h2['display_name']).value
        op = ws2.cell(r, h2['operator']).value
        print(f'  id={nid}  {str(dn):<42}  op={op}')

print('\nDropped IDs (should not appear):')
found_dropped = [ws2.cell(r, h2['id']).value
                 for r in range(3, ws2.max_row + 1)
                 if ws2.cell(r, h2['id']).value in DROP_IDS]
print(f'  {found_dropped if found_dropped else "None — all dropped correctly"}')
