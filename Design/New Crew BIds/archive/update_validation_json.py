#!/usr/bin/env python3
"""
Rebuild validation_json column with rich UI-control schema.

Each parameter slot (A/B/C) becomes an object:
  {
    "type":    data type (see TYPE TAXONOMY below),
    "format":  enforced input format string,
    "label":   UI field label,
    "multi":   true  → comma-separated multi-entry / chip input,
    "options": [...]  → enum only,
    "min":     N      → int only
  }

TYPE TAXONOMY
  time_of_day   clock time 00:00–23:59        format HH:MM
  credit        elapsed credit/block hours     format HH:MM
  duration      long elapsed time (TAFB etc)   format HHH:MM
  int           integer count
  percent_or_time  NN% or HHH:MM (Credit/TAFB ratio)
  date          calendar date                  format MM/DD/YYYY
  date_or_dow   date or day-of-week chip       format MM/DD/YYYY or Mon
  airport       IATA 3-letter code             autocomplete + chips
  pairing       pairing numeric ID             text + chips
  flight        4-digit flight number          text + chips
  crew_id       employee number                text / lookup
  enum          fixed option list              dropdown
"""
from pathlib import Path
import openpyxl, json, copy, datetime
from openpyxl.styles import Alignment

BASE = Path("/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds")

# ── rich validation_json keyed by property_id ────────────────────────────────
VJ = {
    # ── Pairing ──────────────────────────────────────────────────────────────
    1:  {"A": {"type":"airport",          "format":"IATA",       "label":"Airports",       "multi":True}},
    2:  {"A": {"type":"pairing",                                 "label":"Pairing IDs",    "multi":True}},
    3:  {"A": {"type":"time_of_day",      "format":"HH:MM",      "label":"From"},
         "B": {"type":"time_of_day",      "format":"HH:MM",      "label":"To"}},
    4:  {"A": {"type":"credit",           "format":"HH:MM",      "label":"Min Credit"},
         "B": {"type":"credit",           "format":"HH:MM",      "label":"Max Credit"}},
    6:  {"A": {"type":"duration",         "format":"HHH:MM",     "label":"Min TAFB"},
         "B": {"type":"duration",         "format":"HHH:MM",     "label":"Max TAFB"}},
    8:  {"A": {"type":"int",                                     "label":"Min Legs",       "min":1},
         "B": {"type":"int",                                     "label":"Max Legs",       "min":1}},
    9:  {"A": {"type":"credit",           "format":"HH:MM",      "label":"Min Credit"},
         "B": {"type":"credit",           "format":"HH:MM",      "label":"Max Credit"}},
    11: {"A": {"type":"int",                                     "label":"Min Legs",       "min":1},
         "B": {"type":"int",                                     "label":"Max Legs",       "min":1}},
    13: {"A": {"type":"time_of_day",      "format":"HH:MM",      "label":"From"},
         "B": {"type":"time_of_day",      "format":"HH:MM",      "label":"To"}},
    16: {"A": {"type":"int",                                     "label":"Min Days",       "min":1},
         "B": {"type":"int",                                     "label":"Max Days",       "min":1}},
    19: {"A": {"type":"airport",          "format":"IATA",       "label":"Airports",       "multi":True}},
    22: {"A": {"type":"duration",         "format":"HHH:MM",     "label":"Min Duration"},
         "B": {"type":"duration",         "format":"HHH:MM",     "label":"Max Duration"}},
    23: {"A": {"type":"time_of_day",      "format":"HH:MM",      "label":"From"},
         "B": {"type":"time_of_day",      "format":"HH:MM",      "label":"To"}},
    26: {"A": {"type":"flight",           "format":"4-digit",    "label":"Flight Numbers", "multi":True}},
    27: {"A": {"type":"crew_id",                                 "label":"Employee Numbers","multi":True}},
    28: {"A": {"type":"int",                                     "label":"Min Legs",       "min":1},
         "B": {"type":"int",                                     "label":"Max Legs",       "min":1}},
    29: {"A": {"type":"int",                                     "label":"Min",            "min":0},
         "B": {"type":"int",                                     "label":"Max",            "min":0}},
    30: {"A": {"type":"duration",         "format":"HHH:MM",     "label":"Min Duration"},
         "B": {"type":"duration",         "format":"HHH:MM",     "label":"Max Duration"}},
    31: {"A": {"type":"time_of_day",      "format":"HH:MM",      "label":"From"},
         "B": {"type":"time_of_day",      "format":"HH:MM",      "label":"To"}},
    32: {"A": {"type":"date_or_dow",                             "label":"Dates / Days",   "multi":True}},
    33: {"A": {"type":"credit",           "format":"HH:MM",      "label":"Min Block"},
         "B": {"type":"credit",           "format":"HH:MM",      "label":"Max Block"}},
    34: {"A": {"type":"percent_or_time",  "format":"NN or HHH:MM","label":"Credit Rate"}},
    36: {"A": {"type":"credit",           "format":"HH:MM",      "label":"Min Block"},
         "B": {"type":"credit",           "format":"HH:MM",      "label":"Max Block"}},
    37: {"A": {"type":"date_or_dow",                             "label":"Dates / Days",   "multi":True}},
    38: {"A": {"type":"time_of_day",      "format":"HH:MM",      "label":"From"},
         "B": {"type":"time_of_day",      "format":"HH:MM",      "label":"To"}},
    41: {"A": {"type":"duration",         "format":"HHH:MM",     "label":"Min Duration"},
         "B": {"type":"duration",         "format":"HHH:MM",     "label":"Max Duration"}},
    42: {"A": {"type":"int",                                     "label":"Min Legs",       "min":1},
         "B": {"type":"int",                                     "label":"Max Legs",       "min":1}},
    44: {"A": {"type":"airport",          "format":"IATA",       "label":"Airports",       "multi":True}},
    45: {"A": {"type":"int",                                     "label":"Min Legs",       "min":1},
         "B": {"type":"int",                                     "label":"Max Legs",       "min":1}},
    46: {"A": {"type":"credit",           "format":"HH:MM",      "label":"Min Credit"},
         "B": {"type":"credit",           "format":"HH:MM",      "label":"Max Credit"}},
    49: {"A": {"type":"date_or_dow",                             "label":"Dates / Days",   "multi":True}},

    # ── DaysOff ──────────────────────────────────────────────────────────────
    7:  {"A": {"type":"date_or_dow",                             "label":"Dates / Days",   "multi":True},
         "B": {"type":"time_of_day",      "format":"HH:MM",      "label":"Window From"},
         "C": {"type":"time_of_day",      "format":"HH:MM",      "label":"Window To"}},
    12: {"A": {"type":"int",                                     "label":"Min Days",       "min":1},
         "B": {"type":"date",             "format":"MM/DD/YYYY", "label":"Window Start"},
         "C": {"type":"date",             "format":"MM/DD/YYYY", "label":"Window End"}},
    14: {"A": {"type":"int",                                     "label":"Min Days",       "min":1}},
    15: {"A": {"type":"int",                                     "label":"Max Days",       "min":1}},
    21: {"A": {"type":"int",                                     "label":"Min Days Off",   "min":1},
         "B": {"type":"int",                                     "label":"Min Days On",    "min":1},
         "C": {"type":"int",                                     "label":"Max Days On",    "min":1}},
    43: {"A": {"type":"crew_id",                                 "label":"Employee Number"},
         "B": {"type":"int",                                     "label":"Min Shared Days","min":1}},
    48: {"A": {"type":"int",                                     "label":"Max Days",       "min":1}},

    # ── Reserve ──────────────────────────────────────────────────────────────
    5:  {"A": {"type":"enum",                                    "label":"Short Call Type",
               "options":["CRAM","CRPM","PRAM","PRMM","PRPM","RESA","RESB"]}},
    47: {"A": {"type":"date",             "format":"MM/DD/YYYY", "label":"Dates",          "multi":True}},

    # ── Line ─────────────────────────────────────────────────────────────────
    35: {"A": {"type":"int",                                     "label":"Line Number",    "min":1}},
    39: {"A": {"type":"duration",         "format":"HHH:MM",     "label":"Min Duration"}},

    # no params (empty dict) — pids 10,17,18,20,24,25,40
}

# ── load latest file ─────────────────────────────────────────────────────────
files = sorted(BASE.glob("bid_properties-definition-*.xlsx"))
src = files[-1]
print(f"Source: {src.name}")

wb = openpyxl.load_workbook(src)
ws = wb.active

h = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
vj_col  = h['validation_json']
id_col  = h['id']

# Guide row (row 2) — update with schema hint
guide = ws.cell(2, vj_col)
guide.value = '{"A":{"type":"...","format":"...","label":"...","multi":true}}'

updated = 0
for row in range(3, ws.max_row + 1):
    pid_val = ws.cell(row, id_col).value
    try:
        pid = int(pid_val)
    except (TypeError, ValueError):
        continue

    new_vj = VJ.get(pid, {})
    cell = ws.cell(row, vj_col)
    cell.value = json.dumps(new_vj, ensure_ascii=False) if new_vj else '{}'
    # keep existing style; tighten alignment for JSON strings
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    updated += 1

print(f"Updated validation_json for {updated} rows.")

# Widen validation_json column to fit richer content
from openpyxl.utils import get_column_letter
ws.column_dimensions[get_column_letter(vj_col)].width = 55

# ── save ─────────────────────────────────────────────────────────────────────
ts = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f"bid_properties-definition-{ts}.xlsx"
wb.save(out)
print(f"Saved: {out.name}")

# ── spot-check a few rows ────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active
h2  = {ws2.cell(1,c).value: c for c in range(1, ws2.max_column+1)}
spot_pids = [4, 6, 7, 26, 5, 12, 34, 49]
print("\nSpot-check:")
for r in range(3, ws2.max_row+1):
    pid_raw = ws2.cell(r, h2['id']).value
    try:
        pid = int(pid_raw)
    except:
        continue
    if pid in spot_pids:
        tmpl = ws2.cell(r, h2['property_template']).value
        vj   = ws2.cell(r, h2['validation_json']).value
        print(f"  pid={pid:2d}  {str(tmpl):42s}  {vj}")
