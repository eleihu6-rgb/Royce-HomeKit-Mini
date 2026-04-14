#!/usr/bin/env python3
"""
Redesign validation_json to be operator-aware:

  - In-only          → flat {"type":…, "label":…, "multi":true}
  - Has Between      → flat {"type":…, "label":…, "label_from":…, "label_to":…}
  - Single op only   → flat {"type":…, "label":…}   (one param, no label_from/to)
  - Fixed multi-slot → keep A/B/C (different types per slot, e.g. Prefer Off)
  - No params        → {}

UI rule:
  operator=Between  → show two inputs using label_from + label_to
  operator=In       → show multi-value input using label
  other (>,<,=,…)   → show single input using label
"""
import json, datetime, copy
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment

BASE = Path('/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds')
SRC  = BASE / 'bid_properties-definition-2026-03-16-010504.xlsx'

# ── new validation_json keyed by new id ───────────────────────────────────────
# Flat object  = operator-aware single-type slot
# A/B/C object = fixed multi-slot (different param types, structure never changes)

VJ = {
    # ── Pairing ───────────────────────────────────────────────────────────────
    # In-only → flat + multi
    101: {"type": "airport",  "format": "IATA",        "label": "Airports",        "multi": True},
    102: {"type": "pairing",                            "label": "Pairing IDs",     "multi": True},
    104: {"type": "airport",  "format": "IATA",        "label": "Airports",        "multi": True},
    115: {"type": "crew_id",                            "label": "Employee Numbers","multi": True},
    116: {"type": "flight",   "format": "4-digit",     "label": "Flight Numbers",  "multi": True},

    # Boolean / no-param
    117: {},
    128: {},

    # Has Between → flat + label_from + label_to
    103: {"type": "time_of_day", "format": "HH:MM",   "label": "Check-In Time",   "label_from": "From",           "label_to": "To"},
    105: {"type": "credit",      "format": "HH:MM",   "label": "Credit",           "label_from": "Min Credit",     "label_to": "Max Credit"},
    109: {"type": "credit",      "format": "HH:MM",   "label": "Credit",           "label_from": "Min Credit",     "label_to": "Max Credit"},
    111: {"type": "time_of_day", "format": "HH:MM",   "label": "Check-Out Time",  "label_from": "From",           "label_to": "To"},
    112: {"type": "int",                               "label": "Days",             "label_from": "Min Days",       "label_to": "Max Days",      "min": 1},
    113: {"type": "duration",    "format": "HHH:MM",  "label": "TAFB",             "label_from": "Min TAFB",       "label_to": "Max TAFB"},
    114: {"type": "time_of_day", "format": "HH:MM",   "label": "Check-In Time",   "label_from": "From",           "label_to": "To"},
    118: {"type": "duration",    "format": "HHH:MM",  "label": "Duration",         "label_from": "Min Duration",   "label_to": "Max Duration"},
    119: {"type": "duration",    "format": "HHH:MM",  "label": "Duration",         "label_from": "Min Duration",   "label_to": "Max Duration"},
    120: {"type": "time_of_day", "format": "HH:MM",   "label": "Time",             "label_from": "From",           "label_to": "To"},
    122: {"type": "int",                               "label": "Legs",             "label_from": "Min Legs",       "label_to": "Max Legs",      "min": 0},
    126: {"type": "time_of_day", "format": "HH:MM",   "label": "Check-Out Time",  "label_from": "From",           "label_to": "To"},
    127: {"type": "credit",      "format": "HH:MM",   "label": "Block Time",       "label_from": "Min Block Time", "label_to": "Max Block Time"},
    129: {"type": "duration",    "format": "HHH:MM",  "label": "Sit Length",       "label_from": "Min Sit Length", "label_to": "Max Sit Length"},

    # In + Between → flat + label_from/to + multi (UI: In=multi list, Between=two date inputs)
    106: {"type": "date_or_dow",                       "label": "Date / Day",       "label_from": "From Date",      "label_to": "To Date",       "multi": True},
    110: {"type": "date_or_dow",                       "label": "Date / Day",       "label_from": "From Date",      "label_to": "To Date",       "multi": True},
    123: {"type": "date_or_dow",                       "label": "Date / Day",       "label_from": "From Date",      "label_to": "To Date",       "multi": True},

    # Single comparison only (no Between) → flat, one param
    107: {"type": "int",                               "label": "Legs",             "min": 1},   # >, <, =, In
    108: {"type": "int",                               "label": "Legs",             "min": 1},   # =, >, <, In
    121: {"type": "credit",      "format": "HH:MM",   "label": "Block Time"},                   # >, < only
    124: {"type": "int",                               "label": "Legs",             "min": 1},   # In, >, <
    125: {"type": "percent_or_time", "format": "NN or HHH:MM", "label": "Credit Rate"},         # >, <
    130: {"type": "int",                               "label": "Legs",             "min": 1},   # > only
    131: {"type": "airport",     "format": "IATA",    "label": "Airports",         "multi": True},  # op=None, implied In
    132: {"type": "int",                               "label": "Legs",             "min": 0},   # op=None
    133: {"type": "credit",      "format": "HH:MM",   "label": "Credit"},                       # op=None

    # ── DaysOff ───────────────────────────────────────────────────────────────
    # Fixed multi-slot (different types per slot — keep A/B/C)
    201: {"A": {"type": "date_or_dow",                 "label": "Dates / Days",     "multi": True},
          "B": {"type": "time_of_day", "format": "HH:MM", "label": "Window From"},
          "C": {"type": "time_of_day", "format": "HH:MM", "label": "Window To"}},

    204: {"A": {"type": "int",                         "label": "Min Days",         "min": 1},
          "B": {"type": "date",        "format": "MM/DD/YYYY", "label": "Window Start"},
          "C": {"type": "date",        "format": "MM/DD/YYYY", "label": "Window End"}},

    205: {"A": {"type": "int",                         "label": "Min Days Off",     "min": 1},
          "B": {"type": "int",                         "label": "Min Days On",      "min": 1},
          "C": {"type": "int",                         "label": "Max Days On",      "min": 1}},

    206: {"A": {"type": "crew_id",                     "label": "Employee Number"},
          "B": {"type": "int",                         "label": "Min Shared Days",  "min": 1}},

    # Single int
    202: {"type": "int", "label": "Max Days", "min": 1},
    203: {"type": "int", "label": "Min Days", "min": 1},
    207: {"type": "int", "label": "Max Days", "min": 1},

    # ── Reserve ───────────────────────────────────────────────────────────────
    301: {"type": "enum", "label": "Short Call Type",
          "options": ["CRAM", "CRPM", "PRAM", "PRMM", "PRPM", "RESA", "RESB"]},
    302: {"type": "date", "format": "MM/DD/YYYY", "label": "Dates", "multi": True},

    # ── Line ──────────────────────────────────────────────────────────────────
    401: {},
    402: {},
    403: {},
    404: {},
    405: {},
    406: {"type": "int",      "label": "Line Number",  "min": 1},
    407: {"type": "duration", "format": "HHH:MM",      "label": "Min Duration"},
}

# ── load source ───────────────────────────────────────────────────────────────
print(f'Source: {SRC.name}')
wb = openpyxl.load_workbook(SRC)
ws = wb.active
h  = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
vj_col = h['validation_json']
id_col  = h['id']
dn_col  = h['display_name']

# ── update ────────────────────────────────────────────────────────────────────
changed   = []
unchanged = []

for r in range(3, ws.max_row + 1):
    pid = ws.cell(r, id_col).value
    if pid not in VJ:
        unchanged.append(pid)
        continue

    new_val = json.dumps(VJ[pid], ensure_ascii=False)
    cell     = ws.cell(r, vj_col)
    old_val  = cell.value or '{}'

    if new_val != old_val:
        cell.value     = new_val
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        changed.append((pid, ws.cell(r, dn_col).value, old_val, new_val))
    else:
        unchanged.append(pid)

# ── save ──────────────────────────────────────────────────────────────────────
ts  = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f'bid_properties-definition-{ts}.xlsx'
wb.save(out)
print(f'Saved: {out.name}')
print(f'\n{len(changed)} cells updated, {len(unchanged)} unchanged.\n')

for pid, dn, old, new in changed:
    print(f'  id={pid}  {dn}')
    print(f'    OLD: {old}')
    print(f'    NEW: {new}')
    print()
