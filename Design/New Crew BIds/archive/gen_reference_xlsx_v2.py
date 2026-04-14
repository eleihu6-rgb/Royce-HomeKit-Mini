#!/usr/bin/env python3
"""
Generate crew_bids_reference Excel with 2 tabs:
  1. bid_properties      (from bid_properties-definition-Final-Ver2.xlsx)
  2. crew_bids           (parsed from Dec_2025_Bids_Data.xlsx)

Changes from v1:
  - bid_properties tab populated from Final-Ver2 definition (not hardcoded)
  - property_id populated with new IDs (101-407) from definition column A
  - Two new columns after Q (minimum_n):
      R = Legend Property   (N-PBS / legacy bid string pattern)
      S = Remastered Property  (new crew-portal display name)
"""
import re, datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE    = Path("/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds")
DEF_SRC = BASE / "bid_properties-definition-Final-Ver2.xlsx"
DEC_SRC = BASE / "Dec_2025_Bids_Data.xlsx"
PERIOD  = "Dec 2025"

# ── styles ────────────────────────────────────────────────────────────────────
thin   = Side(border_style="thin", color="FFBFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FILL   = PatternFill("solid", fgColor="FF1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFFFF", name="Calibri", size=11)
GUIDE_FILL = PatternFill("solid", fgColor="FFFFF2CC")
GUIDE_FONT = Font(italic=True, name="Calibri", size=10, color="FF595959")
DATA_FONT  = Font(name="Calibri", size=10)

TYPE_FILL = {
    "Pairing": PatternFill("solid", fgColor="FFE2EFDA"),
    "DaysOff": PatternFill("solid", fgColor="FFFCE4D6"),
    "Reserve": PatternFill("solid", fgColor="FFFFF2CC"),
    "Line":    PatternFill("solid", fgColor="FFEDEDED"),
}

def hdr(cell, text):
    cell.value = text
    cell.font  = HDR_FONT;  cell.fill = HDR_FILL;  cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def guide(cell, text):
    cell.value = text
    cell.font  = GUIDE_FONT;  cell.fill = GUIDE_FILL;  cell.border = border
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

def data(cell, value, fill=None, wrap=False, h_align="left"):
    cell.value = value;  cell.font = DATA_FONT;  cell.border = border
    if fill: cell.fill = fill
    cell.alignment = Alignment(horizontal=h_align, vertical="top", wrap_text=wrap)

def col_w(ws, c, w):
    ws.column_dimensions[get_column_letter(c)].width = w

# ── load definition ───────────────────────────────────────────────────────────
def load_definition():
    """Returns:
       def_rows   : list of full row value lists (data rows only, no guide)
       def_cols   : column name list
       new_id_map : orig_pid -> new_id   (built by matching Legend Property)
       lookup     : new_id  -> {'legend': str, 'remastered': str, 'bid_type': str}
    """
    wb = openpyxl.load_workbook(DEF_SRC)
    ws = wb.active
    # Read raw headers then normalise to snake_case for internal use
    raw_hdr  = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    hdr_map  = {(k.lower().replace(' ', '_') if isinstance(k, str) else k): v
                for k, v in raw_hdr.items()}
    def_cols = [c.lower().replace(' ', '_') if isinstance(c, str) else c
                for c in raw_hdr.keys()]

    id_c   = hdr_map['id']
    leg_c  = hdr_map['legend_property']
    rem_c  = hdr_map['remastered_property']
    bt_c   = hdr_map['bid_type']

    lookup   = {}   # new_id -> dict
    def_rows = []

    for r in range(3, ws.max_row + 1):    # row 2 is guide
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        new_id = row[id_c - 1]
        if new_id is None:
            continue
        legend     = row[leg_c - 1] or ''
        remastered = row[rem_c - 1] or ''
        btype      = row[bt_c - 1]  or ''
        lookup[new_id] = {'legend': legend, 'remastered': remastered, 'bid_type': btype}
        def_rows.append(row)

    # Build orig_pid -> new_id via the same PAIRING_PROPS patterns used in parsing
    # Strategy: match the first "If <keyword>" line in Legend Property against patterns
    new_id_map = _build_orig_new_map(lookup)
    return def_rows, def_cols, new_id_map, lookup


# Orig-pid → new-id mapping (derived from generate_crew_bids_sql.py PAIRING_PROPS + non-pairing parsers)
_ORIG_TO_NEW = {
    # Pairing (orig_pid: new_id)
     1:101,  2:102,  3:103, 19:104,  4:105, 32:106,  8:107, 11:108,  9:109,
    37:110, 13:111, 16:112,  6:113, 23:114, 27:115, 26:116, 25:117, 30:118,
    22:119, 31:120, 33:121, 29:122, 49:123, 28:124, 34:125, 38:126, 36:127,
    40:128, 41:129, 42:130,
    # DaysOff
     7:201, 15:202, 14:203, 12:204, 21:205, 43:206, 48:207,
    # Reserve
     5:301, 47:302,
    # Line
    10:401, 17:402, 18:403, 20:404, 24:405, 35:406, 39:407,
}

def _build_orig_new_map(lookup):
    return dict(_ORIG_TO_NEW)   # use the verified hardcoded mapping


# ── Tab 1: bid_properties (from definition) ───────────────────────────────────
def build_bid_properties(wb, def_rows, def_cols):
    ws = wb.create_sheet("bid_properties")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 40

    # Column widths for definition columns
    widths = {
        'id': 7, 'bid_type': 10, 'legend_property': 38, 'remastered_property': 40,
        'award_or_avoid': 18, 'any_or_every': 18, 'operator': 28,
        'validation_json': 55, 'tooltip': 58, 'notes': 42, 'crew_count': 9,
    }

    for c, col_name in enumerate(def_cols, 1):
        hdr(ws.cell(1, c), col_name)
        col_w(ws, c, widths.get(col_name, 14))

    # Guide row (row 2) — use column names as guide text
    GUIDES = {
        'id': 'SMALLINT PK', 'bid_type': 'Pairing|DaysOff|Reserve|Line',
        'Legend Property': 'Actual bid pattern(s) from May+Dec',
        'Remastered Property': 'Crew-facing name on portal',
        'award_or_avoid': '["award"]|["avoid"]|["award","avoid"]|null',
        'any_or_every': '["any"]|["every"]|["any","every"]|null',
        'operator': '["In"]|["Between"]|[">=","<="]|null',
        'validation_json': '{"type":"...","label":"..."}',
        'tooltip': 'Actual bid examples',
        'notes': 'Implementation notes',
        'crew_count': '# unique crews',
    }
    for c, col_name in enumerate(def_cols, 1):
        guide(ws.cell(2, c), GUIDES.get(col_name, ''))

    bt_idx = def_cols.index('bid_type')
    id_idx = def_cols.index('id')

    for r_idx, row in enumerate(def_rows, 3):
        btype = row[bt_idx]
        fill  = TYPE_FILL.get(btype, PatternFill())
        for c, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c, value=val)
            cell.font = DATA_FONT;  cell.fill = fill;  cell.border = border
            if c == id_idx + 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.freeze_panes = "A3"
    print(f"  bid_properties: {len(def_rows)} rows")


# ── Parsers (from generate_crew_bids_sql.py) ──────────────────────────────────
A_E = r'(?:Any|Every)'
PAIRING_PROPS = [
    ( 3, re.compile(r'^Pairing Check-In Time',                  re.I)),
    (13, re.compile(r'^Pairing Check-Out Time',                 re.I)),
    (16, re.compile(r'^Pairing Length',                         re.I)),
    ( 2, re.compile(r'^Pairing Number',                         re.I)),
    (32, re.compile(r'^Departing On',                           re.I)),
    (45, re.compile(r'^' + A_E + r'\s+Duty Legs\s+Counting Deadhead', re.I)),
    ( 8, re.compile(r'^' + A_E + r'\s+Duty Legs',              re.I)),
    ( 1, re.compile(r'^' + A_E + r'\s+Landing In',             re.I)),
    (44, re.compile(r'^' + A_E + r'\s+Duty In',                re.I)),
    (19, re.compile(r'^' + A_E + r'\s+Layover In',             re.I)),
    (30, re.compile(r'^' + A_E + r'\s+Duty Duration',          re.I)),
    ( 6, re.compile(r'^TAFB',                                   re.I)),
    ( 9, re.compile(r'^Average Daily Credit',                   re.I)),
    (46, re.compile(r'^Average Credit',                         re.I)),
    ( 4, re.compile(r'^(?:Pairing\s+)?Total Credit',            re.I)),
    (31, re.compile(r'^' + A_E + r'\s+Duty On Time',           re.I)),
    (37, re.compile(r'^' + A_E + r'\s+Duty On(?:\s+Date)?',   re.I)),
    (28, re.compile(r'^Total Legs In First Duty',               re.I)),
    (42, re.compile(r'^Total Legs In Last Duty',                re.I)),
    (11, re.compile(r'^Total Legs In Pairing',                  re.I)),
    (26, re.compile(r'^' + A_E + r'\s+Flight Number',          re.I)),
    (25, re.compile(r'^' + A_E + r'\s+Leg Is Redeye',          re.I)),
    (34, re.compile(r'^Credit Per Time Away From Base',         re.I)),
    (23, re.compile(r'^' + A_E + r'\s+Enroute Check-In Time',  re.I)),
    (38, re.compile(r'^' + A_E + r'\s+Enroute Check-Out Time', re.I)),
    (22, re.compile(r'^' + A_E + r'\s+Layover Of Duration',    re.I)),
    (49, re.compile(r'^' + A_E + r'\s+Layover On(?:\s+Date)?', re.I)),
    (27, re.compile(r'^' + A_E + r'\s+Leg With Employee Number', re.I)),
    (29, re.compile(r'^Deadhead Legs',                          re.I)),
    (33, re.compile(r'^Average Daily Block Time',               re.I)),
    (36, re.compile(r'^(?:Pairing\s+)?Total Block Time',        re.I)),
    (40, re.compile(r'^Deadhead Day',                           re.I)),
    (41, re.compile(r'^' + A_E + r'\s+Sit Length',             re.I)),
]

DATE_RE = re.compile(
    r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s+\d{4}', re.I)

def _norm_val(v):
    v = str(v).strip()
    v = re.sub(r'\s+(legs?|days?)\s*$', '', v, flags=re.I).strip()
    return v or None

def _norm_list(v):
    parts = [p.strip() for p in str(v).split(',')]
    return ','.join(p for p in parts if p) or None

def _strip_mods(s):
    limit_n = aon = None
    s = re.sub(r'\s*Else Start Next Bid Group\s*$', '', s, flags=re.I).strip()
    m = re.search(r'\s*Limit\s+(\d+)\s*$', s, re.I)
    if m:
        limit_n = int(m.group(1));  s = s[:m.start()].strip()
    if re.search(r'\bAll or Nothing\b', s, re.I):
        aon = 1;  s = re.sub(r'\s*All or Nothing\s*', ' ', s, flags=re.I).strip()
    return s, limit_n, aon

def _extract_op(rem):
    s = rem.strip()
    m = re.match(r'^Between\s+(.+?)\s+And\s+(.+)$', s, re.I)
    if m: return 'Between', _norm_val(m.group(1)), _norm_val(m.group(2)), None
    m = re.match(r'^(>=|<=|>|<|=)\s+(.+)$', s)
    if m: return m.group(1), _norm_val(m.group(2)), None, None
    if s: return None, _norm_list(s), None, None
    return None, None, None, None

def _match_prop(clause):
    clause = clause.strip()
    for pid, pat in PAIRING_PROPS:
        m = pat.match(clause)
        if m:
            return pid, clause[m.end():].strip()
    return None, None

def _parse_prefer_off(s):
    body = re.sub(r'^Prefer Off\s*', '', s, flags=re.I).strip()
    minimum_n = None
    m = re.search(r'\s+Minimum\s+(\d+)\s*$', body, re.I)
    if m:
        minimum_n = int(m.group(1));  body = body[:m.start()].strip()
    pb = pc = None
    tw = re.search(r'\s+Between\s+(\d{1,3}:\d{2})\s+And\s+(\d{1,3}:\d{2})\s*$', body, re.I)
    if tw:
        pb, pc = tw.group(1), tw.group(2);  body = body[:tw.start()].strip()
    if re.match(r'^Between\b', body, re.I):
        m = re.match(r'^Between\s+(.+?)\s+And\s+(.+)$', body, re.I)
        if m: return 'Between', m.group(1).strip(), m.group(2).strip(), pc, minimum_n
    dp = r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s+\d{4}'
    dr = re.match(r'^(' + dp + r')\s*[-\u2013]\s*(' + dp + r')\s*$', body, re.I)
    if dr: return 'Between', dr.group(1).strip(), dr.group(2).strip(), pc, minimum_n
    th = re.match(r'^(.+?)\s+Through\s+(.+)$', body, re.I)
    if th: return 'Between', th.group(1).strip(), th.group(2).strip(), pc, minimum_n
    dates = DATE_RE.findall(body)
    if dates: return None, ','.join(d.strip() for d in dates), pb, pc, minimum_n
    return None, _norm_list(body) if body else None, pb, pc, minimum_n

def _parse_set_cond(s):
    m = re.match(
        r'^Set Condition\s+(\d+)\s+Consecutive Days Off In A Row'
        r'(?:\s+Between\s+(.+?)\s+And\s+(.+))?$', s, re.I)
    if m: return 12, ('Between' if m.group(2) else None), m.group(1), m.group(2), m.group(3)
    m = re.match(r'^Set Condition Minimum Days Off In A Row\s+(\d+)$', s, re.I)
    if m: return 14, None, m.group(1), None, None
    m = re.match(r'^Set Condition Maximum Days Off In A Row\s+(\d+)$', s, re.I)
    if m: return 48, None, m.group(1), None, None
    m = re.match(r'^Set Condition Maximum Days On In A Row\s+(\d+)$', s, re.I)
    if m: return 15, None, m.group(1), None, None
    m = re.match(
        r'^Set Condition Pattern Between\s+(\d+)\s+and\s+(\d+)\s+Days On'
        r',?\s+with\s+(\d+)\s+Days Off', s, re.I)
    if m: return 21, 'Between', m.group(3), m.group(1), m.group(2)
    m = re.match(r'^Set Condition Days Off Opposite Employee\s+(\d+)\s+Minimum\s+(\d+)$', s, re.I)
    if m: return 43, None, m.group(1), m.group(2), None
    if re.match(r'^Set Condition Minimum Credit Window$', s, re.I):   return 18, None, None, None, None
    if re.match(r'^Set Condition Maximum Credit Window$', s, re.I):   return 17, None, None, None, None
    m = re.match(r'^Set Condition Minimum Base Layover\s+(\S+)$', s, re.I)
    if m: return 39, None, m.group(1), None, None
    if re.match(r'^Set Condition No Same Day Pairings$', s, re.I):    return 20, None, None, None, None
    m = re.match(r'^Set Condition Short Call Type\s+(\S+)$', s, re.I)
    if m: return 5, None, m.group(1), None, None
    return None, None, None, None, None

# ── per-row legend derivation ─────────────────────────────────────────────────
# Keyword string to display for each pairing orig_pid (human-readable, matches
# the "If <keyword>" portion of the N-PBS bid string).
_PAIRING_KEYWORD = {
     1: "Any Landing In",       2: "Pairing Number",
     3: "Pairing Check-In Time",4: "Pairing Total Credit",
     6: "TAFB",                  8: "Any/Every Duty Legs",
     9: "Average Daily Credit", 11: "Total Legs In Pairing",
    13: "Pairing Check-Out Time",16: "Pairing Length",
    19: "Any/Every Layover In", 22: "Any/Every Layover Of Duration",
    23: "Any Enroute Check-In Time", 25: "Any Leg Is Redeye",
    26: "Any Flight Number",    27: "Any/Every Leg With Employee Number",
    28: "Total Legs In First Duty", 29: "Deadhead Legs",
    30: "Any/Every Duty Duration", 31: "Any Duty On Time",
    32: "Departing On",          33: "Average Daily Block Time",
    34: "Credit Per Time Away From Base", 36: "Pairing Total Block Time",
    37: "Any/Every Duty On",    38: "Any Enroute Check-Out Time",
    40: "Deadhead Day",          41: "Any/Every Sit Length",
    42: "Total Legs In Last Duty", 49: "Any/Every Layover On",
}

_NEW_TO_ORIG = {v: k for k, v in _ORIG_TO_NEW.items()}

def derive_legend(new_id, action_id, operator, param_a, param_b, param_c,
                  minimum_n, all_or_nothing, lookup):
    """Return the specific N-PBS pattern string for this one crew_bids row."""
    orig_pid = _NEW_TO_ORIG.get(new_id)
    btype    = lookup.get(new_id, {}).get('bid_type', '')

    # ── Pairing rows ──────────────────────────────────────────────────────
    if btype == 'Pairing' and orig_pid:
        kw = _PAIRING_KEYWORD.get(orig_pid, f'property_{orig_pid}')
        if action_id == 1:
            return f'Award Pairings If {kw}'
        if action_id == 2:
            return f'Avoid Pairings If {kw}'
        # node_id > 1 (chained condition — no Award/Avoid prefix)
        return kw

    # ── Prefer Off (201) ─────────────────────────────────────────────────
    if new_id == 201:
        pa = str(param_a or '')
        is_weekends = re.search(r'\bWeekends?\b', pa, re.I)
        if is_weekends:
            if minimum_n is not None and all_or_nothing:
                return f'Prefer Off Weekends Minimum {minimum_n} All or Nothing'
            if minimum_n is not None:
                return f'Prefer Off Weekends Minimum {minimum_n}'
            return 'Prefer Off Weekends'
        return 'Prefer Off'

    # ── Set Condition / non-pairing ───────────────────────────────────────
    if new_id == 202: return 'Set Condition Maximum Days On In A Row'
    if new_id == 203: return 'Set Condition Minimum Days Off In A Row'
    if new_id == 204:
        return ('Set Condition N Consecutive Days Off In A Row Between date1 And date2'
                if operator == 'Between'
                else 'Set Condition N Consecutive Days Off In A Row')
    if new_id == 205: return 'Set Condition Pattern Between A and B Days On, with C Days Off'
    if new_id == 206: return 'Set Condition Days Off Opposite Employee E Minimum N'
    if new_id == 301: return 'Set Condition Short Call Type'
    if new_id == 302: return 'Award Reserve Day On'
    if new_id == 401: return 'Set Condition Maximum Credit Window'
    if new_id == 402: return 'Set Condition Minimum Credit Window'
    if new_id == 403: return 'Clear Schedule and Start Next Bid Group'
    if new_id == 404: return 'Set Condition No Same Day Pairings'
    if new_id == 405: return 'Waive No Same Day Duty Starts'
    if new_id == 406: return 'Forget Line'
    if new_id == 407: return 'Set Condition Minimum Base Layover'

    # Fallback: first line of the definition legend
    full = lookup.get(new_id, {}).get('legend', '')
    return full.split('\n')[0] if full else ''


SKIP = {'Pairing Bid Group', 'Reserve Bid Group'}

def _nd(ai, pi, op, pa, pb, pc, ni, ao, ln, aon, mn):
    return dict(action_id=ai, property_id=pi, operator=op,
                param_a=pa, param_b=pb, param_c=pc,
                node_id=ni, and_or_or=ao,
                limit_n=ln, all_or_nothing=aon, minimum_n=mn)

def parse_bid(raw):
    """Returns list of node dicts, [] to skip, or None on parse error."""
    s, ln, aon = _strip_mods(raw)
    if s in SKIP: return []
    if re.match(r'^Clear Schedule and Start Next Bid Group$', s, re.I):
        return [_nd(None, 10, None, None, None, None, 1, None, ln, aon, None)]
    m = re.match(r'^Forget Line\s+(\d+)$', s, re.I)
    if m: return [_nd(None, 35, None, m.group(1), None, None, 1, None, ln, aon, None)]
    if re.match(r'^Waive No Same Day Duty Starts$', s, re.I):
        return [_nd(None, 24, None, None, None, None, 1, None, ln, aon, None)]
    if re.match(r'^Prefer Off\b', s, re.I):
        op, pa, pb, pc, mn = _parse_prefer_off(s)
        return [_nd(None, 7, op, pa, pb, pc, 1, None, ln, aon, mn)]
    m = re.match(r'^Award Reserve Day On\s+(.+)$', s, re.I)
    if m:
        dates = DATE_RE.findall(m.group(1))
        pa = ','.join(d.strip() for d in dates) if dates else _norm_list(m.group(1))
        return [_nd(None, 47, None, pa, None, None, 1, None, ln, aon, None)]
    if re.match(r'^Set Condition\b', s, re.I):
        pid, op, pa, pb, pc = _parse_set_cond(s)
        if pid: return [_nd(None, pid, op, pa, pb, pc, 1, None, ln, aon, None)]
        return None
    ma = re.match(r'^Award Pairings\s+If\s+', s, re.I)
    mv = re.match(r'^Avoid Pairings\s+If\s+', s, re.I)
    if ma or mv:
        ai   = 1 if ma else 2
        rest = s[(ma or mv).end():]
        nodes = []
        for i, clause in enumerate(re.split(r'\s+If\s+', rest, flags=re.I)):
            clause = clause.strip()
            if not clause: continue
            if re.match(r'^Followed By Pairings', clause, re.I): break
            pid, rem = _match_prop(clause)
            if pid is None: return None
            op, pa, pb, pc = _extract_op(rem)
            nodes.append(_nd(
                ai if i == 0 else None, pid, op, pa, pb, pc,
                i + 1, None if i == 0 else 'AND',
                ln if i == 0 else None,
                aon if i == 0 else None, None))
        return nodes if nodes else None
    return None


# ── Tab 3: crew_bids ──────────────────────────────────────────────────────────
def build_crew_bids(wb, orig_to_new, lookup):
    ws = wb.create_sheet("crew_bids")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 60

    columns = [
        # (header,              width,  guide text)
        ("id",               8,  "BIGINT AUTO_INCREMENT"),
        ("crew_id",         10,  "MEDIUMINT UNSIGNED"),
        ("bid_context",     14,  "Default | Current"),
        ("period",          14,  "e.g. Dec 2025"),
        ("layer",            8,  "1 … 24"),
        ("property_group_id",18, "FK: id of node_id=1 row in same group"),
        ("node_id",         10,  "1 = action row\n2+ = chained condition"),
        ("and_or_or",       10,  "AND | OR\n(NULL for node_id=1)"),
        ("action_id",       12,  "1=Award  2=Avoid\nNULL for non-Pairing"),
        ("property_id",     13,  "→ bid_properties.id\n(new 101-407 IDs)"),
        ("operator",        14,  "> < >= <= =\nBetween | NULL"),
        ("param_a",         18,  "value or lower bound\n(list: comma-separated)"),
        ("param_b",         18,  "upper bound\n(Between only)"),
        ("param_c",         12,  "3rd value\n(if needed)"),
        ("limit_n",         10,  "Limit N\n(max pairings to award)"),
        ("all_or_nothing",  15,  "1 = All or Nothing\nmodifier on Prefer Off"),
        ("minimum_n",       12,  "Minimum N\nfor Prefer Off Weekends"),
        # ── new columns ─────────────────────────────────────────────────────
        ("legend_property",     45, "N-PBS / legacy bid string pattern\n(from definition)"),
        ("remastered_property", 38, "New crew-portal display name\n(from definition)"),
    ]

    for c, (name, width, guide_txt) in enumerate(columns, 1):
        hdr(ws.cell(1, c), name)
        guide(ws.cell(2, c), guide_txt)
        col_w(ws, c, width)

    ws.freeze_panes = "A3"

    # ── parse Dec data ────────────────────────────────────────────────────
    print(f"Parsing {DEC_SRC.name} …")
    src_wb = openpyxl.load_workbook(DEC_SRC, read_only=True, data_only=True)
    src_ws = src_wb["Dec 2025 Bids"]

    errors    = []
    row_num   = 3
    group_id  = 0
    src_count = 0
    written   = 0
    unmapped  = {}   # orig_pid -> count for properties not in mapping

    DA = Alignment(horizontal="left", vertical="top")
    CA = Alignment(horizontal="center", vertical="top")

    for src_row in src_ws.iter_rows(min_row=3, values_only=True):
        bid_str = src_row[5]
        if not bid_str: continue
        bid_str = str(bid_str).strip()
        if not bid_str or bid_str in SKIP: continue

        src_count += 1
        crew_id = str(src_row[1]).strip()
        bid_ctx = str(src_row[2]).strip()
        layer   = int(src_row[3])

        nodes = parse_bid(bid_str)
        if nodes is None:
            errors.append(f"crew={crew_id:>6}  ctx={bid_ctx:<8}  layer={layer}  |  {bid_str}")
            continue
        if not nodes:
            continue

        group_id += 1
        for n in nodes:
            orig_pid = n['property_id']
            new_id   = orig_to_new.get(orig_pid)
            if new_id is None:
                unmapped[orig_pid] = unmapped.get(orig_pid, 0) + 1
                new_id = orig_pid   # fallback: keep orig

            leg_info   = lookup.get(new_id, {})
            legend     = derive_legend(new_id, n['action_id'], n['operator'],
                                       n['param_a'], n['param_b'], n['param_c'],
                                       n['minimum_n'], n['all_or_nothing'], lookup)
            remastered = leg_info.get('remastered', '')
            btype      = leg_info.get('bid_type', '')
            fill       = TYPE_FILL.get(btype)

            row_vals = [
                None,             # id (auto)
                crew_id,          # crew_id
                bid_ctx,          # bid_context
                PERIOD,           # period
                layer,            # layer
                group_id,         # property_group_id
                n['node_id'],     # node_id
                n['and_or_or'],   # and_or_or
                n['action_id'],   # action_id
                new_id,           # property_id ← new ID from definition
                n['operator'],    # operator
                n['param_a'],     # param_a
                n['param_b'],     # param_b
                n['param_c'],     # param_c
                n['limit_n'],     # limit_n
                n['all_or_nothing'], # all_or_nothing
                n['minimum_n'],   # minimum_n
                legend,           # R: Legend Property
                remastered,       # S: Remastered Property
            ]

            for c, val in enumerate(row_vals, 1):
                cell = ws.cell(row_num, c, value=val)
                cell.font = DATA_FONT
                if fill: cell.fill = fill
                # center id/numeric columns
                if c in (1, 5, 6, 7, 9, 10, 15, 16):
                    cell.alignment = CA
                else:
                    cell.alignment = DA

            row_num += 1
            written += 1

    # ── Errors tab ────────────────────────────────────────────────────────
    ews = wb.create_sheet("Errors")
    for c, h_txt in enumerate(["crew_id","bid_context","layer","raw_bid"], 1):
        hdr(ews.cell(1, c), h_txt)
        col_w(ews, c, 12 if c < 4 else 100)
    for r, line in enumerate(errors, 2):
        parts = re.match(
            r'crew=\s*(\S+)\s+ctx=(\S+)\s+layer=(\d+)\s+\|\s+(.*)', line)
        if parts:
            for c, v in enumerate([parts.group(1), parts.group(2),
                                    int(parts.group(3)), parts.group(4)], 1):
                data(ews.cell(r, c), v)
        else:
            data(ews.cell(r, 1), line)

    print(f"  Source rows : {src_count}")
    print(f"  Written rows: {written}")
    print(f"  Groups      : {group_id}")
    print(f"  Errors      : {len(errors)}")
    print(f"  Match rate  : {100*(src_count-len(errors))/src_count:.1f}%")
    if unmapped:
        print(f"  Unmapped orig_pids: {unmapped}")
    return written


# ── Build & save ──────────────────────────────────────────────────────────────
def main():
    print(f"Definition: {DEF_SRC.name}")
    print(f"Dec data  : {DEC_SRC.name}")
    print()

    def_rows, def_cols, orig_to_new, lookup = load_definition()
    print(f"  Definition loaded: {len(def_rows)} properties, {len(lookup)} in lookup")
    print()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_bid_properties(wb, def_rows, def_cols)
    crew_rows = build_crew_bids(wb, orig_to_new, lookup)

    ts  = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
    out = BASE / f'crew_bids_reference-{ts}.xlsx'
    wb.save(out)

    print()
    print(f"Saved: {out.name}")
    print(f"  crew_bids        : {crew_rows} data rows")
    print()

    # ── spot-check columns R/S ────────────────────────────────────────────
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["crew_bids"]
    h2  = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}
    print("Spot-check (first 8 data rows — columns property_id, legend_property, remastered_property):")
    for r in range(3, min(11, ws2.max_row + 1)):
        pid  = ws2.cell(r, h2['property_id']).value
        leg  = (ws2.cell(r, h2['legend_property']).value or '').replace('\n','|')[:55]
        rem  = ws2.cell(r, h2['remastered_property']).value
        print(f"  row{r:3d}  pid={str(pid):<5}  {str(rem):<40}  {leg}")

if __name__ == '__main__':
    main()
