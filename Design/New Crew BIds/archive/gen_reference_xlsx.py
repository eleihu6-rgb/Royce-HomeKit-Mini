#!/usr/bin/env python3
"""
Generate crew_bids_reference.xlsx with 3 tabs:
  1. bid_actions
  2. bid_properties  (colour-coded by bid_type)
  3. crew_bids       (guide row + all 14 459 data rows from crew_bids_insert.sql)
"""
from pathlib import Path
import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE       = Path("/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/DataSource/Crew Bids Design")
OUT        = BASE / "crew_bids_reference.xlsx"
INSERT_SQL = BASE / "crew_bids_insert.sql"

# ── colours ─────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue header
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
GUIDE_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow guide row
GUIDE_FONT = Font(italic=True, name="Calibri", size=10, color="595959")

TYPE_FILL = {
    "Pairing": PatternFill("solid", fgColor="E2EFDA"),  # light green
    "Reserve": PatternFill("solid", fgColor="FFF2CC"),  # light yellow
    "DaysOff": PatternFill("solid", fgColor="FCE4D6"),  # light orange
    "Line":    PatternFill("solid", fgColor="EDEDED"),  # light grey
}

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_style(cell, text):
    cell.value = text
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def data_cell(cell, value, fill=None, wrap=False, align="left"):
    cell.value = value
    cell.font = Font(name="Calibri", size=10)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = border

def guide_cell(cell, text):
    cell.value = text
    cell.font = GUIDE_FONT
    cell.fill = GUIDE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Tab 1: bid_actions ───────────────────────────────────────────────────────
def build_bid_actions(wb):
    ws = wb.create_sheet("bid_actions")
    ws.row_dimensions[1].height = 22
    headers = ["id", "name"]
    for c, h in enumerate(headers, 1):
        hdr_style(ws.cell(1, c), h)
    rows = [(1, "Award"), (2, "Avoid")]
    for r, (rid, name) in enumerate(rows, 2):
        data_cell(ws.cell(r, 1), rid, align="center")
        data_cell(ws.cell(r, 2), name)
    set_col_width(ws, 1, 6)
    set_col_width(ws, 2, 14)

# ── Tab 2: bid_properties ────────────────────────────────────────────────────
def build_bid_properties(wb):
    ws = wb.create_sheet("bid_properties")
    ws.row_dimensions[1].height = 22
    headers = ["id", "bid_type", "property_template", "validation_json", "notes"]
    for c, h in enumerate(headers, 1):
        hdr_style(ws.cell(1, c), h)

    # sorted by usage count (desc) from Dec 2025 import data
    props = [
        ( 1,"Pairing","Landing In A",                        '{"A":"list:airport"}',                          "A = comma-separated IATA airports"),
        ( 2,"Pairing","Pairing Number In A",                 '{"A":"list:pairing"}',                          "A = comma-separated pairing IDs"),
        ( 3,"Pairing","Pairing Check-In Time",               '{"A":"time","B":"time"}',                       None),
        ( 4,"Pairing","Pairing Total Credit",                '{"A":"time","B":"time"}',                       None),
        ( 5,"Reserve","Short Call Type A",                   '{"A":"enum:short_call"}',                       "A = one of: CRAM CRPM PRAM PRMM PRPM RESA RESB"),
        ( 6,"Pairing","TAFB",                                '{"A":"time","B":"time"}',                       None),
        ( 7,"DaysOff","Prefer Off A",                        '{"A":"list:date_or_dow","B":"time","C":"time"}', "A=dates or days | B,C=time window (operator=Between)"),
        ( 8,"Pairing","Duty Legs",                           '{"A":"int","B":"int"}',                         None),
        ( 9,"Pairing","Average Daily Credit",                '{"A":"time","B":"time"}',                       None),
        (10,"Line",   "Clear Schedule and Start Next Bid Group",'{}',                                         None),
        (11,"Pairing","Total Legs In Pairing",               '{"A":"int","B":"int"}',                         None),
        (12,"DaysOff","Min A Consecutive Days Off",          '{"A":"int","B":"date","C":"date"}',              "A=min N days | operator=Between: B=start_date C=end_date (window in which N days must fall)"),
        (13,"Pairing","Pairing Check-Out Time",              '{"A":"time","B":"time"}',                       None),
        (14,"DaysOff","Min Consecutive Days Off A",          '{"A":"int"}',                                   None),
        (15,"DaysOff","Max Consecutive Days On A",           '{"A":"int"}',                                   None),
        (16,"Pairing","Pairing Length",                      '{"A":"int","B":"int"}',                         None),
        (17,"Line",   "Max Credit Window",                   '{}',                                            None),
        (18,"Line",   "Min Credit Window",                   '{}',                                            None),
        (19,"Pairing","Layover In A",                        '{"A":"list:airport"}',                          None),
        (20,"Line",   "No Same Day Pairings",                '{}',                                            None),
        (21,"DaysOff","Min A Days Off Pattern B to C Days On",'{"A":"int","B":"int","C":"int"}',              "A=min days off | B=min days on C=max days on (always operator=Between)"),
        (22,"Pairing","Layover Duration",                    '{"A":"time","B":"time"}',                       None),
        (23,"Pairing","Enroute Check-In Time",               '{"A":"time","B":"time"}',                       None),
        (24,"Line",   "Waive No Same Day Duty Starts",       '{}',                                            None),
        (25,"Pairing","Any Leg Is Redeye",                   '{}',                                            "No param; Counting Deadhead variant noted at import"),
        (26,"Pairing","Any Flight Number A",                 '{"A":"list:flight"}',                           "A = comma-separated flight numbers e.g. 0600,0518"),
        (27,"Pairing","Leg With Employee Number A",          '{"A":"list:crew_id"}',                          "A = comma-separated employee numbers"),
        (28,"Pairing","Total Legs In First Duty",            '{"A":"int","B":"int"}',                         None),
        (29,"Pairing","Deadhead Legs",                       '{"A":"int","B":"int"}',                         None),
        (30,"Pairing","Duty Duration",                       '{"A":"time","B":"time"}',                       None),
        (31,"Pairing","Duty Start Time",                     '{"A":"time","B":"time"}',                       "Time the duty starts"),
        (32,"Pairing","Departing On A",                      '{"A":"list:date_or_dow"}',                      "A = dates or day-of-week names"),
        (33,"Pairing","Average Daily Block Time",            '{"A":"time","B":"time"}',                       None),
        (34,"Pairing","Credit Per Time Away From Base",      '{"A":"percent"}',                               "A = percentage e.g. 70, or time value e.g. 007:00"),
        (35,"Line",   "Forget Line A",                       '{"A":"int"}',                                   "A = line number to forget"),
        (36,"Pairing","Pairing Total Block Time",            '{"A":"time","B":"time"}',                       None),
        (37,"Pairing","Duty On A",                           '{"A":"list:date_or_dow"}',                      "Date/DOW the duty falls on; operator=Between for date range"),
        (38,"Pairing","Enroute Check-Out Time",              '{"A":"time","B":"time"}',                       None),
        (39,"Line",   "Min Base Layover A",                  '{"A":"duration"}',                              "A = HHH:MM e.g. 013:00"),
        (40,"Pairing","Deadhead Day",                        '{}',                                            "Pairing contains a deadhead day"),
        (41,"Pairing","Sit Duration",                        '{"A":"time","B":"time"}',                       None),
        (42,"Pairing","Total Legs In Last Duty",             '{"A":"int","B":"int"}',                         None),
        (43,"DaysOff","Shared Days Off With Employee A Min B Days",'{"A":"crew_id","B":"int"}',              "A=employee ID  B=minimum days"),
        # ── count = 0 in Dec 2025 data (defined, not yet used) ───────────────
        (44,"Pairing","Duty In A",                           '{"A":"list:airport"}',                          None),
        (45,"Pairing","Duty Legs Including Deadhead",        '{"A":"int","B":"int"}',                         None),
        (46,"Pairing","Average Pairing Credit",              '{"A":"time","B":"time"}',                       None),
        (47,"Reserve","Reserve Day On A",                    '{"A":"list:date"}',                             "A = comma-separated dates"),
        (48,"DaysOff","Max Consecutive Days Off A",          '{"A":"int"}',                                   None),
        (49,"Pairing","Layover On A",                        '{"A":"list:date_or_dow"}',                      "Layover On date or DOW; operator=Between for date range"),
    ]

    for r, (pid, btype, tmpl, vj, notes) in enumerate(props, 2):
        fill = TYPE_FILL[btype]
        data_cell(ws.cell(r, 1), pid,   fill=fill, align="center")
        data_cell(ws.cell(r, 2), btype, fill=fill)
        data_cell(ws.cell(r, 3), tmpl,  fill=fill)
        data_cell(ws.cell(r, 4), vj,    fill=fill)
        data_cell(ws.cell(r, 5), notes, fill=fill, wrap=True)

    set_col_width(ws, 1,  5)
    set_col_width(ws, 2, 10)
    set_col_width(ws, 3, 38)
    set_col_width(ws, 4, 30)
    set_col_width(ws, 5, 50)
    ws.freeze_panes = "A2"

# ── Tab 3: crew_bids ─────────────────────────────────────────────────────────
def build_crew_bids(wb):  # returns row count
    ws = wb.create_sheet("crew_bids")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 60

    columns = [
        ("id",               8,  "BIGINT AUTO_INCREMENT"),
        ("crew_id",          10, "MEDIUMINT UNSIGNED"),
        ("bid_context",      14, "Default | Current"),
        ("period",           14, "e.g. Dec 2025"),
        ("layer",            8,  "1 … 24"),
        ("property_group_id",18, "= id of node_id=1 row in same group"),
        ("node_id",          10, "1 = action row\n2+ = chained condition"),
        ("and_or_or",        10, "AND | OR\n(NULL for node_id=1)"),
        ("action_id",        12, "1=Award  2=Avoid\nNULL for non-Pairing"),
        ("property_id",      13, "→ bid_properties.id"),
        ("operator",         14, "> < >= <= =\nBetween | NULL"),
        ("param_a",          16, "value or lower bound\n(list: comma-separated)"),
        ("param_b",          16, "upper bound\n(Between only)"),
        ("param_c",          12, "3rd value\n(if needed)"),
        ("limit_n",          10, "Limit N\n(max pairings to award)"),
        ("all_or_nothing",   15, "1 = All or Nothing\nmodifier on Prefer Off"),
        ("minimum_n",        12, "Minimum N\nfor Prefer Off Weekends"),
    ]

    for c, (name, width, guide) in enumerate(columns, 1):
        hdr_style(ws.cell(1, c), name)
        guide_cell(ws.cell(2, c), guide)
        set_col_width(ws, c, width)

    ws.freeze_panes = "A3"

    # ── parse + write data rows from INSERT SQL ──────────────────────────────
    print("Parsing crew_bids_insert.sql …")

    # Tokenise a SQL values tuple — handles strings with embedded parens/commas
    def parse_row(line):
        """Return list of 16 Python values from one SQL values line, or None."""
        # Strip leading/trailing whitespace, parens, trailing comma
        line = line.strip().rstrip(',').rstrip(';')
        if not (line.startswith('(') and line.endswith(')')):
            return None
        inner = line[1:-1]

        tokens = []
        i = 0
        while i < len(inner):
            if inner[i] == ' ':
                i += 1
                continue
            if inner[i] == "'":
                # quoted string — scan to closing quote (handle '' escapes)
                j = i + 1
                while j < len(inner):
                    if inner[j] == "'" and (j + 1 >= len(inner) or inner[j+1] != "'"):
                        break
                    if inner[j] == "'" and inner[j+1] == "'":
                        j += 2
                        continue
                    j += 1
                tokens.append(inner[i+1:j].replace("''", "'"))
                i = j + 1
                if i < len(inner) and inner[i] == ',':
                    i += 1
            else:
                # unquoted token (NULL or integer)
                j = inner.find(',', i)
                if j == -1:
                    j = len(inner)
                tok = inner[i:j].strip()
                tokens.append(None if tok == 'NULL' else int(tok))
                i = j + 1

        return tokens if len(tokens) == 16 else None

    row_num = 3   # Excel row (1=header, 2=guide, 3+ = data)
    count = 0
    DATA_FONT  = Font(name="Calibri", size=10)
    DATA_ALIGN = Alignment(horizontal="left", vertical="top")

    for line in INSERT_SQL.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped.startswith('('):
            continue
        row = parse_row(stripped)
        if row is None:
            continue
        # First token is crew_id — must be an integer
        if not isinstance(row[0], int):
            continue
        row_data = [None] + row   # id column left blank (DB assigns it)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row_num, c, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
        row_num += 1
        count += 1

    return count

# ── Build & save ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default blank sheet

build_bid_actions(wb)
build_bid_properties(wb)
crew_rows = build_crew_bids(wb)

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  bid_actions    : 2 rows")
print(f"  bid_properties : 50 rows")
print(f"  crew_bids      : {crew_rows} data rows")
