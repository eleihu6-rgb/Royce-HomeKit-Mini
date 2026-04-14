#!/usr/bin/env python3
"""
gen_final_v3.py

Step 1: Collect tooltip examples from 4-month crew bid TXT files.
Step 2: Build orig_pid -> new_id mapping from xlsx.
Step 3: Format tooltip text per new_id.
Step 4: Generate new Excel with tooltip column.
"""
import sys, re, datetime
from pathlib import Path
from collections import defaultdict

# ── Import server module ───────────────────────────────────────────────────────
if 'server' in sys.modules:
    del sys.modules['server']
sys.path.insert(0, str(Path('/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit')))
import server as srv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path('/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds')
SRC  = BASE / 'bid_properties-definition-2026-03-16-062922.xlsx'

TXT_FILES = [
    BASE / 'May Crew Bids 2025 All in one.txt',
    BASE / 'Jul 2025 All in one.txt',
    BASE / 'Aug 2025 All in one.txt',
    BASE / 'Dec Crew Bids 2025 All in one.txt',
]

# ── DN_TO_ORIG: display_name -> orig_pid for non-pairing rows ─────────────────
DN_TO_ORIG = {
    "Prefer Off": 7, "Max Consecutive Days On": 15, "Min Consecutive Days Off": 14,
    "Min Consecutive Days Off In Window": 12, "Days Off / Days On Pattern": 21,
    "Shared Days Off With Employee": 43, "Short Call Type": 5, "Reserve Day On": 47,
    "Max Credit Window": 17, "Min Credit Window": 18,
    "Clear Schedule and Start Next Bid Group": 10, "No Same Day Pairings": 20,
    "Waive No Same Day Duty Starts": 24, "Forget Line": 35, "Min Base Layover": 39,
}

# ── Operator sort order for pairing tooltips ──────────────────────────────────
OP_ORDER = {'In': 0, '=': 1, '>': 2, '<': 3, 'Between': 4}

# ==============================================================================
# STEP 1 — Collect tooltip examples from 4 months
# ==============================================================================
print('Step 1: Parsing TXT files...')

# pairing_examples[orig_pid][(action, op_key, any_every_flag)] = full_bid_str (first seen)
pairing_examples = defaultdict(dict)

# non_pairing_examples[orig_pid] = list of unique bid strings (up to 5)
non_pairing_examples = defaultdict(list)
non_pairing_seen = defaultdict(set)

QUALIFIER_SKIP = re.compile(r'\(PairingStage|\(Counting Deadhead|\(Ordered', re.I)

for txt_path in TXT_FILES:
    print(f'  Parsing: {txt_path.name}')
    if not txt_path.exists():
        print(f'    WARNING: File not found, skipping.')
        continue
    content = txt_path.read_text(encoding='utf-8', errors='replace')
    # Detect period from filename (just pass filename stem)
    rows, errors, stats = srv.nb_parse_txt(content, txt_path.stem)
    print(f'    -> {len(rows)} rows, {len(errors)} errors')

    for nd in rows:
        bid_type = None
        orig_pid = nd.get('property_id')
        raw_bid  = nd.get('raw_bid', '')

        if not raw_bid or orig_pid is None:
            continue

        s, _, _ = srv._nb_strip_modifiers(raw_bid)
        s_short  = s[:120]

        # Determine bid_type by action_id and property_id
        # Pairing bids start with Award/Avoid Pairings If
        if re.match(r'^(Award|Avoid)\s+Pairings\s+If\b', s, re.I):
            # Pairing property
            action_match = re.match(r'^(Award|Avoid)\s+Pairings\s+If\s+', s, re.I)
            if not action_match:
                continue
            action = action_match.group(1).capitalize()  # "Award" or "Avoid"
            rest   = s[action_match.end():]

            # Split on \s+If\s+ to get clauses
            clauses = re.split(r'\s+If\s+', rest, flags=re.I)
            for clause in clauses:
                clause = clause.strip()
                if not clause:
                    continue
                # Skip qualifier-modified variants
                if QUALIFIER_SKIP.search(clause):
                    continue

                cpid, remainder = srv._nb_match_pairing_prop(clause)
                if cpid is None or remainder is None:
                    continue

                op, pa, pb, pc = srv._nb_extract_op_params(remainder)

                # Determine op_key
                if op is None and pa is not None:
                    op_key = 'In'
                elif op is not None:
                    op_key = op
                else:
                    op_key = ''

                # any_every_flag
                if re.search(r'\bAny\b', clause, re.I):
                    any_every_flag = 'Any'
                elif re.search(r'\bEvery\b', clause, re.I):
                    any_every_flag = 'Every'
                else:
                    any_every_flag = ''

                dedup_key = (action, op_key, any_every_flag)
                if dedup_key not in pairing_examples[cpid]:
                    pairing_examples[cpid][dedup_key] = s_short

        elif re.match(r'^(Prefer Off|Set Condition|Award Reserve Day On|'
                      r'Clear Schedule and Start Next Bid Group|Forget Line|'
                      r'Waive No Same Day Duty Starts)', s, re.I):
            # Non-pairing property
            if s not in non_pairing_seen[orig_pid]:
                if len(non_pairing_examples[orig_pid]) < 5:
                    non_pairing_examples[orig_pid].append(s_short)
                    non_pairing_seen[orig_pid].add(s)

print(f'  Pairing pids with examples: {sorted(pairing_examples.keys())}')
print(f'  Non-pairing pids with examples: {sorted(non_pairing_examples.keys())}')

# ==============================================================================
# STEP 2 — Build orig_pid -> new_id mapping from xlsx
# ==============================================================================
print('\nStep 2: Building orig_pid -> new_id mapping...')

wb_src = openpyxl.load_workbook(SRC)
ws_src = wb_src.active
h = {ws_src.cell(1, c).value: c for c in range(1, ws_src.max_column + 1)}

# new_id -> orig_pid map
new_id_to_orig_pid = {}
# new_id -> bid_type
new_id_to_btype = {}
# new_id -> all source row values
src_rows = []  # list of (new_id, vals_list)

cols = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]
guide_row = [ws_src.cell(2, c).value for c in range(1, ws_src.max_column + 1)]

for r in range(3, ws_src.max_row + 1):
    new_id = ws_src.cell(r, h['id']).value
    bt     = ws_src.cell(r, h['bid_type']).value
    dn     = ws_src.cell(r, h['display_name']).value
    npbs   = ws_src.cell(r, h['N-PBS Property']).value or ''

    vals = [ws_src.cell(r, c).value for c in range(1, ws_src.max_column + 1)]
    src_rows.append((new_id, vals))
    new_id_to_btype[new_id] = bt

    if bt == 'Pairing':
        # Use first line of N-PBS Property to match
        first_line = npbs.split('\n')[0].strip()
        # Extract the clause after "Award/Avoid Pairings If "
        m = re.match(r'^(?:Award|Avoid)\s+Pairings\s+If\s+', first_line, re.I)
        if m:
            clause = first_line[m.end():].strip()
            cpid, _ = srv._nb_match_pairing_prop(clause)
            if cpid is not None:
                new_id_to_orig_pid[new_id] = cpid
                print(f'  Pairing id={new_id} -> orig_pid={cpid}  ({dn})')
            else:
                print(f'  Pairing id={new_id} -> NO MATCH for clause: {clause!r}')
        else:
            print(f'  Pairing id={new_id} -> could not parse N-PBS: {first_line!r}')
    else:
        # Non-pairing: use DN_TO_ORIG
        if dn in DN_TO_ORIG:
            new_id_to_orig_pid[new_id] = DN_TO_ORIG[dn]
            print(f'  Non-pairing id={new_id} -> orig_pid={DN_TO_ORIG[dn]}  ({dn})')
        else:
            print(f'  Non-pairing id={new_id} -> NO MATCH for dn={dn!r}')

# ==============================================================================
# STEP 3 — Format tooltip text per new_id
# ==============================================================================
print('\nStep 3: Building tooltips...')

tooltips = {}  # new_id -> tooltip string

for new_id, vals in src_rows:
    bt = new_id_to_btype.get(new_id)
    orig_pid = new_id_to_orig_pid.get(new_id)

    if orig_pid is None:
        tooltips[new_id] = ''
        continue

    if bt == 'Pairing':
        examples_dict = pairing_examples.get(orig_pid, {})
        if not examples_dict:
            tooltips[new_id] = ''
            continue

        # Sort: Award first, then by operator order
        def sort_key(item):
            (action, op_key, any_every_flag), bid_str = item
            action_order = 0 if action == 'Award' else 1
            op_order = OP_ORDER.get(op_key, 99)
            return (action_order, op_order, any_every_flag)

        sorted_examples = sorted(examples_dict.items(), key=sort_key)
        lines = [bid_str for _, bid_str in sorted_examples[:10]]
        tooltips[new_id] = '\n'.join(lines)

    else:
        # Non-pairing
        examples = non_pairing_examples.get(orig_pid, [])
        tooltips[new_id] = '\n'.join(examples[:5])

# ==============================================================================
# STEP 4 — Generate new Excel
# ==============================================================================
print('\nStep 4: Building Excel...')

# ── Style constants ────────────────────────────────────────────────────────────
thin   = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
GUIDE_FONT  = Font(italic=True, name='Calibri', size=10, color='595959')
GUIDE_FILL  = PatternFill('solid', fgColor='FFF2CC')
GUIDE_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
TYPE_FILL = {
    'Pairing': PatternFill('solid', fgColor='E2EFDA'),
    'DaysOff': PatternFill('solid', fgColor='FCE4D6'),
    'Reserve': PatternFill('solid', fgColor='FFF2CC'),
    'Line':    PatternFill('solid', fgColor='EDEDED'),
}
DATA_FONT  = Font(name='Calibri', size=10)
DATA_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
ID_ALIGN   = Alignment(horizontal='center', vertical='top')
COL_WIDTHS = {
    'id': 7, 'bid_type': 10, 'Legend Property': 38, 'Remastered Property': 40,
    'award_or_avoid': 18, 'any_or_every': 18, 'operator': 28,
    'validation_json': 58, 'tooltip': 60, 'notes': 42, 'crew_count': 9,
}

# ── Build new column list ──────────────────────────────────────────────────────
# Rename: N-PBS Property -> Legend Property, display_name -> Remastered Property
# Insert tooltip after validation_json
new_cols = []
for col_name in cols:
    if col_name == 'N-PBS Property':
        new_cols.append('Legend Property')
    elif col_name == 'display_name':
        new_cols.append('Remastered Property')
    else:
        new_cols.append(col_name)
    if col_name == 'validation_json':
        new_cols.append('tooltip')

print(f'  New columns: {new_cols}')

# Map old col index -> new col name for data values
# We need to know where tooltip inserts in the new_cols list
tooltip_new_idx = new_cols.index('tooltip')  # 0-based

# Build guide row for new columns
old_guide_by_name = {cols[i]: guide_row[i] for i in range(len(cols))}
new_guide_row = []
for col_name in new_cols:
    if col_name == 'tooltip':
        new_guide_row.append('Actual bid examples from May/Jul/Aug/Dec data')
    elif col_name == 'Legend Property':
        new_guide_row.append(old_guide_by_name.get('N-PBS Property', ''))
    elif col_name == 'Remastered Property':
        new_guide_row.append(old_guide_by_name.get('display_name', ''))
    else:
        new_guide_row.append(old_guide_by_name.get(col_name, ''))

# ── Create workbook ────────────────────────────────────────────────────────────
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = 'bid_properties'

# Row 1 — header
for c, col_name in enumerate(new_cols, 1):
    cell = ws.cell(1, c, value=col_name)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = HDR_ALIGN
    cell.border    = border
ws.row_dimensions[1].height = 22

# Row 2 — guide
for c, val in enumerate(new_guide_row, 1):
    cell = ws.cell(2, c, value=val)
    cell.font      = GUIDE_FONT
    cell.fill      = GUIDE_FILL
    cell.alignment = GUIDE_ALIGN
    cell.border    = border
ws.row_dimensions[2].height = 40

# Find column indices (0-based) in old cols for building data rows
old_id_idx  = cols.index('id')
old_bt_idx  = cols.index('bid_type')

# For each new column name, find the corresponding old value source
def get_old_col_name(new_col_name):
    if new_col_name == 'Legend Property':
        return 'N-PBS Property'
    if new_col_name == 'Remastered Property':
        return 'display_name'
    if new_col_name == 'tooltip':
        return None  # special
    return new_col_name

# Data rows
id_new_idx = new_cols.index('id') + 1  # 1-based col for id

for r_idx, (new_id, old_vals) in enumerate(src_rows, 3):
    old_vals_by_name = {cols[i]: old_vals[i] for i in range(len(cols))}
    btype = old_vals_by_name.get('bid_type', '')
    fill  = TYPE_FILL.get(btype, PatternFill())

    for c, col_name in enumerate(new_cols, 1):
        if col_name == 'tooltip':
            val = tooltips.get(new_id, '')
        else:
            old_name = get_old_col_name(col_name)
            val = old_vals_by_name.get(old_name, None)

        cell = ws.cell(r_idx, c, value=val)
        cell.font   = DATA_FONT
        cell.fill   = fill
        cell.border = border
        if col_name == 'id':
            cell.alignment = ID_ALIGN
        else:
            cell.alignment = DATA_ALIGN

# Column widths
for c, col_name in enumerate(new_cols, 1):
    ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS.get(col_name, 14)

ws.freeze_panes = 'A3'

# ── Save ───────────────────────────────────────────────────────────────────────
ts  = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f'bid_properties-definition-{ts}.xlsx'
wb.save(out)
print(f'\nSaved: {out.name}')

# ── Verification ───────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active
h2  = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}

print(f'\nColumn list: {list(h2.keys())}')
print(f'Total data rows: {ws2.max_row - 2}')

print('\nSpot-check tooltips:')
for r in range(3, ws2.max_row + 1):
    nid = ws2.cell(r, h2['id']).value
    if nid in (105, 122, 201):
        tt = ws2.cell(r, h2['tooltip']).value or ''
        print(f'\n  id={nid}:')
        for line in tt.split('\n')[:5]:
            print(f'    {line}')
        if tt.count('\n') >= 5:
            print(f'    ... ({tt.count(chr(10))+1} lines total)')
