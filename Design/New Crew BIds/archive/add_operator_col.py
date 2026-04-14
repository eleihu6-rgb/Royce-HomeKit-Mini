#!/usr/bin/env python3
"""
1. Parse May + Dec TXT files using server.py parsing functions.
2. Collect unique operators actually used per property_id.
3. Insert an 'operator' column (JSON array) into the latest
   bid_properties-definition xlsx, after 'any_or_every' (col E),
   making it the new col F.

Operator JSON format:
  ["In"]                    list-only (no explicit operator token)
  ["Between"]               range only
  [">=","<=",">","<","="]  comparison operators
  ["In","Between",">=",…]  mixed, sorted by frequency descending
  null                      no operator (no params, or fixed param)
"""
import sys, json, copy, datetime, re
from pathlib import Path
from collections import defaultdict

# ── import server parsing functions ──────────────────────────────────────────
REPO = Path("/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit")
sys.path.insert(0, str(REPO))
import server as srv   # safe — server only starts under __main__

TXT_FILES = [
    REPO / "Design/New Crew BIds/May Crew Bids 2025 All in one.txt",
    REPO / "Design/New Crew BIds/Dec Crew Bids 2025 All in one.txt",
]
BASE = REPO / "Design/New Crew BIds"

# ── parse TXT files → collect operators per property_id ──────────────────────
# op_counts[pid][op] = count
op_counts = defaultdict(lambda: defaultdict(int))

for txt_path in TXT_FILES:
    print(f"Parsing {txt_path.name} …")
    txt_content = txt_path.read_text(encoding='utf-8', errors='replace')
    period = srv._nb_detect_period(txt_content)
    out_rows, err_rows, stats = srv.nb_parse_txt(txt_content, period)
    print(f"  → {stats['input_rows']} input rows, {len(out_rows)} parsed, {len(err_rows)} errors")
    for nd in out_rows:
        pid = nd.get('property_id')
        op  = nd.get('operator')
        if pid is None:
            continue
        # Represent "no explicit operator but has param_a" as "In"
        if op is None and nd.get('param_a') is not None:
            op_key = 'In'
        elif op is None:
            op_key = None   # truly no param
        else:
            op_key = op     # Between / >= / <= / > / < / =
        if op_key is not None:
            op_counts[pid][op_key] += 1

print(f"\nOperators found for {len(op_counts)} properties:")
for pid in sorted(op_counts.keys()):
    ops = dict(sorted(op_counts[pid].items(), key=lambda x: -x[1]))
    print(f"  pid={pid:2d}  {ops}")

# ── build operator JSON per pid ───────────────────────────────────────────────
# Order: most-frequent first; fixed canonical ordering within equals comparisons
COMP_ORDER = ['=', '>=', '<=', '>', '<']

def make_op_json(pid):
    counts = op_counts.get(pid)
    if not counts:
        return None   # no operators (no-param properties)
    # Sort by count desc, then canonical order for ties
    def sort_key(op):
        freq = -counts[op]
        pos  = COMP_ORDER.index(op) if op in COMP_ORDER else -1
        return (freq, pos)
    ops = sorted(counts.keys(), key=sort_key)
    return json.dumps(ops)

# ── load latest Excel that does NOT already have 'operator' column ────────────
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

files = sorted(BASE.glob("bid_properties-definition-*.xlsx"))
src = None
for f in reversed(files):
    wb_chk = openpyxl.load_workbook(f, read_only=True)
    ws_chk = wb_chk.active
    existing_cols = [ws_chk.cell(1, c).value for c in range(1, ws_chk.max_column + 1)]
    wb_chk.close()
    if 'operator' not in existing_cols:
        src = f
        break
if src is None:
    raise FileNotFoundError("All definition files already have 'operator' column — delete the latest and re-run.")
print(f"\nSource: {src.name}")

wb = openpyxl.load_workbook(src)
ws = wb.active

h = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
print("Existing columns:", list(h.keys()))

ae_col     = h['any_or_every']    # col E = 5
insert_at  = ae_col + 1           # col F = 6  ← new 'operator' column
id_col     = h['id']

print(f"Inserting 'operator' at column {get_column_letter(insert_at)} ({insert_at})")

# ── shift columns F onwards right by 1 ───────────────────────────────────────
def clone_style(src_cell, dst_cell):
    dst_cell.font      = copy.copy(src_cell.font)
    dst_cell.fill      = copy.copy(src_cell.fill)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.border    = copy.copy(src_cell.border)

max_col = ws.max_column
for col in range(max_col, insert_at - 1, -1):
    dst_col = col + 1
    for row in range(1, ws.max_row + 1):
        sc = ws.cell(row, col)
        dc = ws.cell(row, dst_col)
        dc.value = sc.value
        clone_style(sc, dc)
    src_ltr = get_column_letter(col)
    dst_ltr = get_column_letter(dst_col)
    if src_ltr in ws.column_dimensions:
        ws.column_dimensions[dst_ltr].width = ws.column_dimensions[src_ltr].width

# clear vacated column
for row in range(1, ws.max_row + 1):
    ws.cell(row, insert_at).value = None

# ── fill new 'operator' column ────────────────────────────────────────────────
# borrow styles from adjacent column
def style_from(row, ref_col):
    clone_style(ws.cell(row, ref_col), ws.cell(row, insert_at))

# row 1 header
hdr = ws.cell(1, insert_at)
hdr.value = 'operator'
style_from(1, ae_col)

# row 2 guide
guide = ws.cell(2, insert_at)
guide.value = '["In"] | ["Between"] | [">=","<="] | null'
style_from(2, ae_col)

# data rows
for row in range(3, ws.max_row + 1):
    pid_raw = ws.cell(row, id_col).value
    try:
        pid = int(pid_raw)
    except (TypeError, ValueError):
        continue
    val = make_op_json(pid)
    cell = ws.cell(row, insert_at)
    cell.value = val
    style_from(row, ae_col)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

ws.column_dimensions[get_column_letter(insert_at)].width = 28

# ── save ──────────────────────────────────────────────────────────────────────
ts = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f"bid_properties-definition-{ts}.xlsx"
wb.save(out)
print(f"\nSaved: {out.name}")

# ── print final column order ──────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active
cols = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
print("Final columns:", cols)

# spot-check
h2 = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}
print("\nOperator values (non-null):")
for r in range(3, ws2.max_row + 1):
    v = ws2.cell(r, h2['operator']).value
    if v:
        tmpl = ws2.cell(r, h2['property_template']).value
        pid  = ws2.cell(r, h2['id']).value
        print(f"  pid={str(pid):3s}  {str(tmpl):42s}  {v}")
