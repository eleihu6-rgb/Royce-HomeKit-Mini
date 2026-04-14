#!/usr/bin/env python3
"""
Update column C (N-PBS Property) to include Award/Avoid Pairings If prefix
where applicable, based on actual May+Dec text data.

Format per pairing property cell (newline-separated):
  Award Pairings If Any Landing In
  Award Pairings If Every Landing In     ← if Every also used
  Avoid Pairings If Any Landing In
  Avoid Pairings If Every Landing In     ← if Every also used

Non-pairing properties (Set Condition, Prefer Off, etc.) kept as-is.
"""
import sys, re, datetime, copy
from pathlib import Path
from collections import defaultdict, OrderedDict

if 'server' in sys.modules:
    del sys.modules['server']
sys.path.insert(0, str(Path('/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit')))
import server as srv
import openpyxl
from openpyxl.styles import Alignment

BASE = Path('/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds')
TXT  = [BASE / 'May Crew Bids 2025 All in one.txt',
        BASE / 'Dec Crew Bids 2025 All in one.txt']

# ── collect distinct prop keywords per (pid, action) ─────────────────────────
# action: 'Award' | 'Avoid' | None (Set Condition / Prefer Off etc.)
award_kws = defaultdict(set)   # pid → {keyword, ...}
avoid_kws = defaultdict(set)
other_kws = defaultdict(set)   # pid → {full non-pairing string, ...}

def get_kw(clause, pid):
    """Return matched keyword portion (regex match end only, no params)."""
    for ppid, pat in srv._PAIRING_PROPS:
        if ppid == pid:
            m = pat.match(clause.strip())
            if m:
                return clause.strip()[:m.end()].strip()
    return None

for f in TXT:
    txt = f.read_text(encoding='utf-8', errors='replace')
    out_rows, _, _ = srv.nb_parse_txt(txt, srv._nb_detect_period(txt))
    for nd in out_rows:
        pid = nd.get('property_id')
        raw = nd.get('raw_bid', '')
        if pid is None:
            continue
        s, _, _ = srv._nb_strip_modifiers(raw)

        ma = re.match(r'^Award Pairings\s+If\s+', s, re.I)
        mv = re.match(r'^Avoid Pairings\s+If\s+', s, re.I)

        if ma or mv:
            is_award = bool(ma)
            rest = s[(ma or mv).end():]
            for clause in re.split(r'\s+If\s+', rest, flags=re.I):
                cpid, _ = srv._nb_match_pairing_prop(clause.strip())
                if cpid == pid:
                    kw = get_kw(clause, pid)
                    if kw:
                        (award_kws if is_award else avoid_kws)[pid].add(kw)
        else:
            # Non-pairing: store canonical string (strip trailing param values)
            prop = re.split(
                r'\s+(Between\b|>=|<=|>|<|=\s|\d)', s, maxsplit=1, flags=re.I
            )[0].strip()
            other_kws[pid].add(prop)

# ── build final N-PBS cell value per pid ─────────────────────────────────────
def build_npbs(pid):
    """Return new N-PBS cell value for pairing properties only (Award/Avoid).
    Returns None for non-pairing pids (leave cell unchanged).
    """
    a_kws = sorted(award_kws[pid])
    v_kws = sorted(avoid_kws[pid])

    if not a_kws and not v_kws:
        return None  # non-pairing or not seen — leave unchanged

    lines = []
    for kw in a_kws:
        lines.append(f'Award Pairings If {kw}')
    for kw in v_kws:
        lines.append(f'Avoid Pairings If {kw}')
    return '\n'.join(lines)

# ── load the correct source (with fixed operator for pid=108, clean non-pairing) ──
src = BASE / 'bid_properties-definition-2026-03-16-005404.xlsx'
print(f'Source: {src.name}')

wb = openpyxl.load_workbook(src)
ws = wb.active
h  = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
id_col   = h['id']
npbs_col = h['N-PBS Property']

# ── build new_id → orig_pid mapping ──────────────────────────────────────────
# Read crew_count per new_id, match to orig pid by cross-referencing with
# known crew_count from parsed data.
# Strategy: for each row, try to match current N-PBS content to a known pid
# via the award/avoid/other keyword sets.

def infer_orig_pid(cur_npbs, crew_count):
    cur_parts = set(p.strip() for p in cur_npbs.replace('\n','|').split('|') if p.strip())
    best_pid, best_score = None, 0
    for pid in sorted(set(list(award_kws)+list(avoid_kws)+list(other_kws))):
        all_kws = award_kws[pid] | avoid_kws[pid] | other_kws[pid]
        score = sum(1 for k in all_kws if any(
            k.lower() in c.lower() or c.lower() in k.lower()
            for c in cur_parts
        ))
        if score > best_score:
            best_score, best_pid = score, pid
    return best_pid if best_score > 0 else None

# ── update column C ───────────────────────────────────────────────────────────
changes   = []
unchanged = []
skipped   = []

btype_col = h['bid_type']

for r in range(3, ws.max_row + 1):
    new_id   = ws.cell(r, id_col).value
    btype    = ws.cell(r, btype_col).value or ''
    cur_npbs = ws.cell(r, npbs_col).value or ''

    # Only update Pairing rows — non-pairing N-PBS strings are already complete
    if btype != 'Pairing':
        unchanged.append(new_id)
        continue

    if '— not used' in cur_npbs or cur_npbs == '':
        skipped.append((new_id, cur_npbs))
        continue

    orig_pid = infer_orig_pid(cur_npbs, None)
    if orig_pid is None:
        skipped.append((new_id, cur_npbs))
        continue

    new_val = build_npbs(orig_pid)
    if new_val is None:
        skipped.append((new_id, cur_npbs))
        continue

    cell = ws.cell(r, npbs_col)
    if new_val != cur_npbs:
        cell.value = new_val
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        changes.append((new_id, cur_npbs, new_val))
    else:
        unchanged.append(new_id)

# ── save ──────────────────────────────────────────────────────────────────────
ts  = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f'bid_properties-definition-{ts}.xlsx'
wb.save(out)
print(f'Saved: {out.name}')
print(f'\n{len(changes)} cells updated, {len(unchanged)} unchanged, {len(skipped)} skipped.\n')

for new_id, old, new in changes:
    old_fmt = old.replace('\n', ' | ')
    new_fmt = new.replace('\n', ' | ')
    print(f'  id={new_id:3}  WAS: {old_fmt[:60]}')
    print(f'         NOW: {new_fmt[:90]}')
    print()
