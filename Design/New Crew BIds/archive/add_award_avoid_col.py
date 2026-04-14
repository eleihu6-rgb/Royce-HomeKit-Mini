#!/usr/bin/env python3
"""
Insert column E 'award_or_avoid' (JSON) after column D 'display_name'.

Values (based on actual May+Dec usage):
  ["award", "avoid"]  – property used with both in data
  ["award"]           – only used with Award
  ["avoid"]           – only used with Avoid
  null                – not applicable (DaysOff / Reserve / Line — no award/avoid concept)
"""
import sys, re, json, copy, datetime
from pathlib import Path
from collections import defaultdict

if 'server' in sys.modules:
    del sys.modules['server']
sys.path.insert(0, str(Path('/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit')))
import server as srv
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

BASE = Path('/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds')
TXT  = [BASE / 'May Crew Bids 2025 All in one.txt',
        BASE / 'Dec Crew Bids 2025 All in one.txt']

# ── collect award / avoid usage per orig pid ─────────────────────────────────
award_pids = set()
avoid_pids = set()

def get_kw(clause, pid):
    for ppid, pat in srv._PAIRING_PROPS:
        if ppid == pid:
            m = pat.match(clause.strip())
            if m:
                return clause.strip()[:m.end()].strip()
    return None

for f in TXT:
    txt = f.read_text(encoding='utf-8', errors='replace')
    out_rows, _, _ = srv.nb_parse_txt(txt, srv._nb_detect_period(txt))
    for nd in out_rows:
        pid = nd.get('property_id')
        raw = nd.get('raw_bid', '')
        if pid is None:
            continue
        s, _, _ = srv._nb_strip_modifiers(raw)
        ma = re.match(r'^Award Pairings\s+If\s+', s, re.I)
        mv = re.match(r'^Avoid Pairings\s+If\s+', s, re.I)
        if ma or mv:
            is_award = bool(ma)
            rest = s[(ma or mv).end():]
            for clause in re.split(r'\s+If\s+', rest, flags=re.I):
                cpid, _ = srv._nb_match_pairing_prop(clause.strip())
                if cpid == pid:
                    if is_award:
                        award_pids.add(pid)
                    else:
                        avoid_pids.add(pid)

def make_award_avoid_json(orig_pid):
    has_award = orig_pid in award_pids
    has_avoid = orig_pid in avoid_pids
    if has_award and has_avoid:
        return '["award", "avoid"]'
    if has_award:
        return '["award"]'
    if has_avoid:
        return '["avoid"]'
    return None   # non-pairing

print('Award/Avoid per orig pid:')
all_pairing_pids = sorted(award_pids | avoid_pids)
for pid in all_pairing_pids:
    print(f'  pid={pid:2d}  {make_award_avoid_json(pid)}')

# ── load latest file ──────────────────────────────────────────────────────────
src = BASE / 'bid_properties-definition-2026-03-16-010104.xlsx'
print(f'\nSource: {src.name}')

wb = openpyxl.load_workbook(src)
ws = wb.active

h = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
disp_col   = h['display_name']   # col D = 4
insert_at  = disp_col + 1        # col E = 5
id_col     = h['id']
btype_col  = h['bid_type']

print(f"Inserting 'award_or_avoid' at column {get_column_letter(insert_at)} ({insert_at})")

# ── build new_id → orig_pid via N-PBS Property content matching ───────────────
npbs_col = h['N-PBS Property']

def infer_orig_pid(cur_npbs, btype):
    if btype != 'Pairing':
        return None
    cur_parts = set(p.strip() for p in cur_npbs.replace('\n', '|').split('|') if p.strip())
    best_pid, best_score = None, 0
    for pid in sorted(award_pids | avoid_pids):
        # build keyword set for this pid using get_kw
        score = sum(1 for p in cur_parts
                    if re.search(r'(Award|Avoid)\s+Pairings\s+If\s+', p, re.I)
                    and get_kw(re.sub(r'^(Award|Avoid)\s+Pairings\s+If\s+', '', p, flags=re.I).strip(), pid))
        if score > best_score:
            best_score, best_pid = score, pid
    return best_pid if best_score > 0 else None

# ── shift columns right to make room ─────────────────────────────────────────
def clone_style(src_cell, dst_cell):
    dst_cell.font      = copy.copy(src_cell.font)
    dst_cell.fill      = copy.copy(src_cell.fill)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.border    = copy.copy(src_cell.border)

max_col = ws.max_column
for col in range(max_col, insert_at - 1, -1):
    dst = col + 1
    for row in range(1, ws.max_row + 1):
        sc = ws.cell(row, col)
        dc = ws.cell(row, dst)
        dc.value = sc.value
        clone_style(sc, dc)
    sl = get_column_letter(col)
    dl = get_column_letter(dst)
    if sl in ws.column_dimensions:
        ws.column_dimensions[dl].width = ws.column_dimensions[sl].width

for row in range(1, ws.max_row + 1):
    ws.cell(row, insert_at).value = None

# ── header ────────────────────────────────────────────────────────────────────
hdr = ws.cell(1, insert_at)
hdr.value = 'award_or_avoid'
clone_style(ws.cell(1, disp_col), hdr)

# ── guide row ─────────────────────────────────────────────────────────────────
guide = ws.cell(2, insert_at)
guide.value = '["award"] | ["avoid"] | ["award","avoid"] | null'
clone_style(ws.cell(2, disp_col), guide)

# ── data rows ─────────────────────────────────────────────────────────────────
filled = 0
for row in range(3, ws.max_row + 1):
    btype    = ws.cell(row, btype_col).value or ''
    cur_npbs = ws.cell(row, npbs_col).value or ''
    cell     = ws.cell(row, insert_at)

    clone_style(ws.cell(row, disp_col), cell)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

    if btype != 'Pairing' or '— not used' in cur_npbs:
        cell.value = None
        continue

    orig_pid = infer_orig_pid(cur_npbs, btype)
    val = make_award_avoid_json(orig_pid) if orig_pid else None
    cell.value = val
    if val:
        filled += 1

ws.column_dimensions[get_column_letter(insert_at)].width = 22

# ── save ──────────────────────────────────────────────────────────────────────
ts  = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f'bid_properties-definition-{ts}.xlsx'
wb.save(out)
print(f'Saved: {out.name}  ({filled} pairing rows filled)')

# ── spot-check ────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active
h2  = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}
print(f'\nFinal columns: {list(h2.keys())}')
print('\naward_or_avoid values:')
for r in range(3, ws2.max_row + 1):
    v = ws2.cell(r, h2['award_or_avoid']).value
    if v:
        pid  = ws2.cell(r, h2['id']).value
        dn   = ws2.cell(r, h2['display_name']).value
        print(f'  id={str(pid):3s}  {str(dn):42s}  {v}')
