#!/usr/bin/env python3
"""
Insert column E 'any_or_every' (JSON) after column D 'property_template'.

JSON values:
  ["any", "every"]  – property seen with both Any and Every in N-PBS data
  ["any"]           – property seen with Any only
  ["every"]         – property seen with Every only
  null              – property has no Any/Every qualifier
"""
from pathlib import Path
import openpyxl, re, datetime, copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path("/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds")

files = sorted(BASE.glob("bid_properties-definition-*.xlsx"))
if not files:
    raise FileNotFoundError("No bid_properties-definition-*.xlsx found")
src = files[-1]
print(f"Source: {src.name}")

wb = openpyxl.load_workbook(src)
ws = wb.active

# ── locate columns ────────────────────────────────────────────────────────────
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
npbs_col = headers['N-PBS Property']   # C = 3
tmpl_col = headers['property_template']  # D = 4
insert_at = tmpl_col + 1               # E = 5  ← new column goes here

print(f"Inserting 'any_or_every' at column {get_column_letter(insert_at)} ({insert_at})")

# ── compute any_or_every value per row ───────────────────────────────────────
def classify(npbs_val):
    if not npbs_val:
        return None
    entries = [e.strip() for e in str(npbs_val).splitlines() if e.strip()]
    has_any   = any(re.match(r'^Any\b',   e, re.I) for e in entries)
    has_every = any(re.match(r'^Every\b', e, re.I) for e in entries)
    if has_any and has_every:
        return '["any", "every"]'
    if has_any:
        return '["any"]'
    if has_every:
        return '["every"]'
    return None

# ── style helpers: clone a cell's style ──────────────────────────────────────
def clone_style(src_cell, dst_cell):
    dst_cell.font      = copy.copy(src_cell.font)
    dst_cell.fill      = copy.copy(src_cell.fill)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.border    = copy.copy(src_cell.border)

# ── insert blank column at insert_at by shifting everything right ─────────────
# Must go right-to-left to avoid overwriting
max_col = ws.max_column
for col in range(max_col, insert_at - 1, -1):
    dst_col = col + 1
    for row in range(1, ws.max_row + 1):
        src_cell = ws.cell(row, col)
        dst_cell = ws.cell(row, dst_col)
        dst_cell.value = src_cell.value
        clone_style(src_cell, dst_cell)
    # Copy column width
    src_letter = get_column_letter(col)
    dst_letter = get_column_letter(dst_col)
    if src_letter in ws.column_dimensions:
        ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width

# ── clear the now-vacated column insert_at ───────────────────────────────────
for row in range(1, ws.max_row + 1):
    ws.cell(row, insert_at).value = None

# ── fill new column ───────────────────────────────────────────────────────────
# Borrow header style from column D header
def apply_header_style(cell):
    clone_style(ws.cell(1, tmpl_col), cell)

def apply_guide_style(cell):
    clone_style(ws.cell(2, tmpl_col), cell)

def apply_data_style(cell, tmpl_col_for_row, row):
    clone_style(ws.cell(row, tmpl_col_for_row), cell)

# Row 1: header
hdr_cell = ws.cell(1, insert_at)
hdr_cell.value = 'any_or_every'
apply_header_style(hdr_cell)

# Row 2: guide
guide_cell = ws.cell(2, insert_at)
guide_cell.value = '["any"] | ["every"] | ["any","every"] | null'
apply_guide_style(guide_cell)

# Row 3+: data
for row in range(3, ws.max_row + 1):
    npbs_val = ws.cell(row, npbs_col).value
    val = classify(npbs_val)

    data_cell = ws.cell(row, insert_at)
    data_cell.value = val  # None → empty cell (null)
    apply_data_style(data_cell, tmpl_col, row)
    # JSON cells: use monospace-style alignment
    data_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

# Set column width for new column
ws.column_dimensions[get_column_letter(insert_at)].width = 20

# ── save ──────────────────────────────────────────────────────────────────────
ts = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f"bid_properties-definition-{ts}.xlsx"
wb.save(out)
print(f"Saved: {out.name}")

# ── summary ───────────────────────────────────────────────────────────────────
print("\nany_or_every values assigned:")
wb2 = openpyxl.load_workbook(out)
ws2 = wb2.active
hdrs2 = {ws2.cell(1,c).value: c for c in range(1, ws2.max_column+1)}
ae_col = hdrs2['any_or_every']
tmpl_col2 = hdrs2['property_template']
counts = {}
for row in range(3, ws2.max_row + 1):
    v = ws2.cell(row, ae_col).value or 'null'
    counts[v] = counts.get(v, 0) + 1
    if v != 'null':
        print(f"  row {row:2d}  {ws2.cell(row,1).value!s:8s}  {str(ws2.cell(row,tmpl_col2).value):40s}  {v}")
print(f"\nSummary: {counts}")
