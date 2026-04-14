#!/usr/bin/env python3
"""
1. Rename column D header: property_template → display_name
2. Update display_name values: follow N-PBS names, replace A/B/C letters,
   keep Any/Every prefix, only change where ambiguous.
"""
from pathlib import Path
import openpyxl, datetime, copy
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

BASE = Path("/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds")

# ── display names keyed by property_id ───────────────────────────────────────
# Unchanged entries are still listed so this is the single source of truth.
DISPLAY = {
    # ── Pairing ──────────────────────────────────────────────────────────────
     1: "Any Landing In Airport",
     2: "Pairing Number",
     3: "Pairing Check-In Time",
     4: "Pairing Total Credit",
     6: "TAFB",
     8: "Any/Every Duty Legs",
     9: "Average Daily Credit",
    11: "Total Legs In Pairing",
    13: "Pairing Check-Out Time",
    16: "Pairing Length",
    19: "Any/Every Layover In Airport",
    22: "Any/Every Layover Duration",
    23: "Any Enroute Check-In Time",
    25: "Any Leg Is Redeye",
    26: "Any Flight Number",
    27: "Any/Every Leg With Employee Number",
    28: "Total Legs In First Duty",
    29: "Deadhead Legs",
    30: "Any/Every Duty Duration",
    31: "Any Duty On Time",
    32: "Departing On",
    33: "Average Daily Block Time",
    34: "Credit Per Time Away From Base",
    36: "Pairing Total Block Time",
    37: "Any/Every Duty On Date / Day",
    38: "Any Enroute Check-Out Time",
    40: "Deadhead Day",
    41: "Any/Every Sit Length",
    42: "Total Legs In Last Duty",
    44: "Any/Every Duty In Airport",
    45: "Any/Every Duty Legs Counting Deadhead",
    46: "Average Pairing Credit",
    49: "Any/Every Layover On Date / Day",
    # ── DaysOff ──────────────────────────────────────────────────────────────
     7: "Prefer Off",
    12: "Min Consecutive Days Off In Window",
    14: "Min Consecutive Days Off",
    15: "Max Consecutive Days On",
    21: "Days Off / Days On Pattern",
    43: "Shared Days Off With Employee",
    48: "Max Consecutive Days Off",
    # ── Reserve ──────────────────────────────────────────────────────────────
     5: "Short Call Type",
    47: "Reserve Day On",
    # ── Line ─────────────────────────────────────────────────────────────────
    10: "Clear Schedule and Start Next Bid Group",
    17: "Max Credit Window",
    18: "Min Credit Window",
    20: "No Same Day Pairings",
    24: "Waive No Same Day Duty Starts",
    35: "Forget Line",
    39: "Min Base Layover",
}

# ── load latest file ──────────────────────────────────────────────────────────
files = sorted(BASE.glob("bid_properties-definition-*.xlsx"))
src = files[-1]
print(f"Source: {src.name}")

wb = openpyxl.load_workbook(src)
ws = wb.active

h = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
print("Columns:", list(h.keys()))

tmpl_col = h['property_template']
id_col   = h['id']

# ── rename column header ──────────────────────────────────────────────────────
ws.cell(1, tmpl_col).value = 'display_name'
print(f"\nRenamed column {get_column_letter(tmpl_col)} header: property_template → display_name")

# ── update guide row ──────────────────────────────────────────────────────────
ws.cell(2, tmpl_col).value = 'Crew-facing property name shown on portal'

# ── update data rows ──────────────────────────────────────────────────────────
changed = []
unchanged = []

for row in range(3, ws.max_row + 1):
    pid_raw = ws.cell(row, id_col).value
    try:
        pid = int(pid_raw)
    except (TypeError, ValueError):
        continue

    cell = ws.cell(row, tmpl_col)
    old  = cell.value
    new  = DISPLAY.get(pid)

    if new is None:
        unchanged.append((pid, old))
        continue

    if new != old:
        cell.value = new
        changed.append((pid, old, new))
    else:
        unchanged.append((pid, old))

# ── widen column D slightly to fit longer names ───────────────────────────────
ws.column_dimensions[get_column_letter(tmpl_col)].width = 40

# ── save ──────────────────────────────────────────────────────────────────────
ts = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f"bid_properties-definition-{ts}.xlsx"
wb.save(out)
print(f"Saved: {out.name}")

# ── report ────────────────────────────────────────────────────────────────────
print(f"\n{len(changed)} display names updated:")
for pid, old, new in changed:
    print(f"  pid={str(pid):3s}  '{old}'  →  '{new}'")

print(f"\n{len(unchanged)} unchanged.")
