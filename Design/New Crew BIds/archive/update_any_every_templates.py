#!/usr/bin/env python3
"""
Update bid_properties-definition xlsx:
  If a row's 'N-PBS Property' column has both 'Any X' AND 'Every X' entries,
  rename property_template to use 'Any/Every X …' prefix.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, re

BASE = Path("/Users/kimi/Library/Mobile Documents/com~apple~CloudDocs/DevOps/Royce-HomeKit/Design/New Crew BIds")

# Find latest definition file
files = sorted(BASE.glob("bid_properties-definition-*.xlsx"))
if not files:
    raise FileNotFoundError("No bid_properties-definition-*.xlsx found")
src = files[-1]
print(f"Source: {src.name}")

wb = openpyxl.load_workbook(src)
ws = wb.active

# Map header names → column index
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
print("Columns:", list(headers.keys()))

npbs_col = headers.get('N-PBS Property')
tmpl_col = headers.get('property_template')

if not npbs_col or not tmpl_col:
    raise ValueError(f"Expected 'N-PBS Property' and 'property_template' columns. Found: {list(headers.keys())}")

changes = []

for row in range(2, ws.max_row + 1):
    npbs_cell = ws.cell(row, npbs_col)
    tmpl_cell = ws.cell(row, tmpl_col)
    npbs_val = npbs_cell.value
    tmpl_val = tmpl_cell.value

    if not npbs_val or not tmpl_val:
        continue

    # N-PBS Property entries are newline-separated within the cell
    entries = [e.strip() for e in str(npbs_val).splitlines() if e.strip()]

    has_any   = any(re.match(r'^Any\b',   e, re.I) for e in entries)
    has_every = any(re.match(r'^Every\b', e, re.I) for e in entries)

    if has_any and has_every:
        # Strip leading Any / Every / Any/Every from existing template
        clean = re.sub(r'^(Any/Every|Any|Every)\s+', '', str(tmpl_val), flags=re.I).strip()
        new_tmpl = f"Any/Every {clean}"

        if new_tmpl != tmpl_val:
            changes.append((row, tmpl_val, new_tmpl))
            tmpl_cell.value = new_tmpl
            # Preserve existing cell style (font/fill already set), just update value

if changes:
    print(f"\n{len(changes)} property_template(s) updated:")
    for row, old, new in changes:
        print(f"  row {row}: '{old}' → '{new}'")
else:
    print("\nNo changes needed (no properties with both Any + Every in N-PBS Property).")

# Save with new timestamp
ts = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
out = BASE / f"bid_properties-definition-{ts}.xlsx"
wb.save(out)
print(f"\nSaved: {out.name}")
