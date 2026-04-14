#!/usr/bin/env python3
"""
NAVOPS — PDF Roster Parser Backend
Run: python3 server.py
Serves the HTML on :8080 and the parse API on POST /api/convert
"""
import http.server, json, os, re, io, traceback, subprocess, threading, uuid, time, base64, queue
import concurrent.futures
from collections import defaultdict
from urllib.parse import urlparse, parse_qs

ANALYSE_WORKERS = 4
TABLE_LIMIT     = 50

# ─── Live-reload file watcher ────────────────────────────────────────────────
_lr_clients  = set()   # set of queue.Queue, one per connected browser tab
_lr_lock     = threading.Lock()

def _lr_broadcast():
    with _lr_lock:
        for q in list(_lr_clients):
            try:
                q.put_nowait('reload')
            except queue.Full:
                pass

def _lr_watch():
    """Poll UI files every 400 ms; broadcast 'reload' on any mtime change."""
    BASE = os.path.dirname(os.path.abspath(__file__))
    patterns = [
        [BASE],                                          # *.html at root
        [BASE, 'css'],                                   # css/*.css
        [BASE, 'js'],                                    # js/*.js
        [BASE, 'pages'],                                 # pages/*.html
    ]
    exts = {'.html', '.css', '.js'}
    mtimes = {}

    def collect():
        for parts in patterns:
            d = os.path.join(*parts)
            if not os.path.isdir(d):
                continue
            for fname in os.listdir(d):
                if os.path.splitext(fname)[1] in exts:
                    fp = os.path.join(d, fname)
                    try:
                        mtimes[fp] = os.stat(fp).st_mtime
                    except OSError:
                        pass

    collect()
    while True:
        time.sleep(0.4)
        changed = False
        for parts in patterns:
            d = os.path.join(*parts)
            if not os.path.isdir(d):
                continue
            for fname in os.listdir(d):
                if os.path.splitext(fname)[1] not in exts:
                    continue
                fp = os.path.join(d, fname)
                try:
                    mt = os.stat(fp).st_mtime
                except OSError:
                    continue
                if mtimes.get(fp) != mt:
                    mtimes[fp] = mt
                    changed = True
        if changed:
            _lr_broadcast()

threading.Thread(target=_lr_watch, daemon=True, name='livereload-watcher').start()

# ─── Background job store ───────────────────────────────────────────────────────
_jobs = {}   # job_id -> {'status': 'running'|'done'|'error', 'tables': [], 'progress': N, 'error': str}

SQL_DATA_DIR   = os.path.expanduser('~/rois_tg_live_load')
MYSQL_CMD      = ['mysql', '-uroot', '-pR@iscrew2026', 'rois_tg_live_prod']
WHITELIST_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'table_whitelist.md')

try:
    import pdfplumber
except ImportError:
    raise SystemExit("pdfplumber not installed. Run: pip3 install pdfplumber --break-system-packages")

# ─── Parser ───────────────────────────────────────────────────────────────────
DUTY_PATTERN = re.compile(
    r'FLYXX|FLY|SIM|VGDO|TGDO|LEAVE|UNION|WILD|COD|PRPM|PRAM|CRM|CBT|TGS|SCM|V\d{4}|GT|ST|IB|NB|VAC'
)
TIME_RE = re.compile(r'^\d{1,2}:\d{2}$')

def split_codes(text):
    parts = DUTY_PATTERN.findall(text)
    return parts if parts else [text]

def build_col_map(words):
    col = {}
    for w in words:
        if re.fullmatch(r'\d{1,2}', w['text']) and 100 <= w['top'] <= 115:
            d = int(w['text'])
            if 1 <= d <= 31:
                col[d] = (w['x0'] + w['x1']) / 2
        if w['text'] == 'C/IN'  and 95 <= w['top'] <= 115: col['CIN']  = (w['x0']+w['x1'])/2
        if w['text'] == 'C/Out' and 95 <= w['top'] <= 115: col['COUT'] = (w['x0']+w['x1'])/2
    return col

def sorted_int_cols(col_map):
    return sorted(k for k in col_map if isinstance(k, int))

def nearest_col(x, col_map, tol=18):
    best, dist = None, 1e9
    for k, cx in col_map.items():
        d = abs(x - cx)
        if d < dist: dist, best = d, k
    return best if dist < tol else None

def col_dict_split(row_words, col_map):
    """Assign duty codes to columns, splitting merged strings (e.g. UNIONUNIONV4011)."""
    result = {}
    scols = sorted_int_cols(col_map)
    for w in row_words:
        cx    = (w['x0'] + w['x1']) / 2
        parts = split_codes(w['text'])
        if len(parts) == 1:
            c = nearest_col(cx, col_map)
            if c is not None:
                result[c] = result.get(c, '') + parts[0]
        else:
            # Multi-part: assign consecutive columns starting from left edge
            start = nearest_col(w['x0'] + 2, col_map)
            if start is None or start not in scols:
                start = nearest_col(cx, col_map)
            if start in scols:
                idx = scols.index(start)
                for p in parts:
                    if idx < len(scols):
                        result[scols[idx]] = p
                        idx += 1
            elif start == 'CIN':
                result['CIN'] = parts[0]
                idx = 0
                for p in parts[1:]:
                    if idx < len(scols):
                        result[scols[idx]] = p
                        idx += 1
    return result

def time_col_dict(row_words, col_map):
    """Like plain_col_dict but only accepts HH:MM time strings — filters header/footer noise."""
    d = {}
    for w in row_words:
        if not TIME_RE.match(w['text']): continue
        c = nearest_col((w['x0']+w['x1'])/2, col_map)
        if c is not None:
            d[c] = d.get(c, '') + w['text']
    return d

def plain_col_dict(row_words, col_map):
    d = {}
    for w in row_words:
        c = nearest_col((w['x0']+w['x1'])/2, col_map)
        if c is not None:
            d[c] = d.get(c, '') + w['text']
    return d

def infer_period(words):
    """Extract year and month from 'Period: May 2025' header."""
    MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
              'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
    texts = [w['text'].lower() for w in words if w['top'] < 80]
    for i, t in enumerate(texts):
        if t in MONTHS and i+1 < len(texts) and texts[i+1].isdigit():
            return int(texts[i+1]), MONTHS[t]
    return 2025, 5  # fallback

def infer_category(words):
    """Extract base and rank from 'Category: YVR-737-CA' header (e.g. base='YVR', rank='CA')."""
    header = [w for w in words if w['top'] < 80]
    for i, w in enumerate(header):
        if w['text'] == 'Category:' and i + 1 < len(header):
            parts = header[i + 1]['text'].split('-')
            base = parts[0] if parts else ''
            rank = parts[-1] if len(parts) > 1 else ''
            return base, rank
    return '', ''

def parse_pdf_bytes(pdf_bytes):
    """Parse a PDF (as bytes) and return list of duty dicts."""
    duties = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pg_idx, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=2, y_tolerance=2)
            col_map = build_col_map(words)

            if pg_idx == 0:
                year, month = infer_period(words)
                crew_base, crew_rank = infer_category(words)

            rows_by_top = defaultdict(list)
            for w in words:
                rows_by_top[round(w['top'])].append(w)

            crew_starts = []
            for top in sorted(rows_by_top.keys()):
                rw   = sorted(rows_by_top[top], key=lambda w: w['x0'])
                left = [w for w in rw if w['x0'] < 90]
                lt   = [w['text'] for w in left]
                if (len(lt) >= 3 and re.fullmatch(r'\d+', lt[0])
                        and lt[1] == '/' and re.fullmatch(r'\d+', lt[2])):
                    crew_starts.append(top)

            for ci, block_top in enumerate(crew_starts):
                block_end = crew_starts[ci+1]-1 if ci+1 < len(crew_starts) else int(page.height)
                bw = [w for w in words if block_top-8 <= w['top'] <= block_end]

                id_row = sorted([w for w in bw if abs(w['top']-block_top) <= 1 and w['x0'] < 90],
                                key=lambda w: w['x0'])
                it = [w['text'] for w in id_row]
                try:
                    si = it.index('/'); seniority = it[si-1]; crew_id = it[si+1]
                except:
                    continue

                def left_word(offset, tol=2):
                    ws = [w for w in bw if abs(w['top']-block_top-offset) <= tol and w['x0'] < 90]
                    return ws[0]['text'] if ws else ''

                def left_last(offset, tol=2):
                    ws = sorted([w for w in bw if abs(w['top']-block_top-offset) <= tol and w['x0'] < 90],
                                key=lambda w: w['x0'])
                    return ws[-1]['text'] if ws else ''

                name     = left_word(6)
                rp_cr    = left_word(12).replace('Cr:', '')

                # Keyword-based: find "Days Off: <N>" regardless of vertical position
                left_seq = sorted([w for w in bw if w['x0'] < 90], key=lambda w: (w['top'], w['x0']))
                days_off = ''
                for i, w in enumerate(left_seq):
                    if w['text'] == 'Off:' and i + 1 < len(left_seq):
                        days_off = left_seq[i + 1]['text']
                        break

                def get_row(offset, tol=1.5):
                    return [w for w in bw if abs(w['top']-block_top-offset) <= tol and w['x0'] > 90]

                cin_map    = plain_col_dict(get_row(1),      col_map)
                cout_map   = plain_col_dict(get_row(7),      col_map)
                code_map   = col_dict_split(get_row(19, 2),  col_map)
                crhour_map = col_dict_split(get_row(13, 2),  col_map)
                rank_map   = plain_col_dict(get_row(25, 2),  col_map)

                # ── Carry-over from previous month ──
                carry_time = cin_map.get('CIN', '')
                carry_code = code_map.get('CIN', '')
                if re.match(r'\d{1,2}:\d{2}', carry_time) and carry_code:
                    cout_label = cout_map.get('CIN', '')
                    m = re.match(r'([A-Za-z]+)(\d+)', cout_label)
                    prev_day  = int(m.group(2)) if m else 30
                    prev_month = month - 1 if month > 1 else 12
                    prev_year  = year if month > 1 else year - 1
                    end_time   = cout_map.get(1, '')
                    duties.append({
                        'crewId': crew_id, 'seniority': seniority, 'name': name,
                        'crew_base': crew_base, 'crew_rank': crew_rank,
                        'credit': rp_cr, 'daysOff': days_off,
                        'date': f'{prev_year}-{prev_month:02d}-{prev_day:02d}',
                        'startDate': f'{prev_year}-{prev_month:02d}-{prev_day:02d} {carry_time}',
                        'endDate': (f'{year}-{month:02d}-01 {end_time}' if end_time
                                    else f'{prev_year}-{prev_month:02d}-{prev_day:02d}'),
                        'assignment': carry_code,
                        'creditHours': crhour_map.get('CIN', ''),
                        'pairingLabel': carry_code,
                        'actingRank': rank_map.get('CIN', ''),
                    })

                # ── Normal days 1-31 ──
                all_days = set()
                for k in list(code_map.keys()) + list(cin_map.keys()):
                    if isinstance(k, int):
                        all_days.add(k)

                for day in sorted(all_days):
                    code   = code_map.get(day, '')
                    c_in   = cin_map.get(day, '')
                    c_out  = cout_map.get(day, '')
                    crhour   = crhour_map.get(day, '')
                    act_rank = rank_map.get(day, '')
                    if not code and not c_in:
                        continue
                    date_str = f'{year}-{month:02d}-{day:02d}'
                    # Overnight duty: end time spills into next day's C/OUT column
                    if not c_out and day + 1 not in all_days and cout_map.get(day + 1):
                        next_cout = cout_map[day + 1]
                        if day < 31:
                            end_str = f'{year}-{month:02d}-{day+1:02d} {next_cout}'
                        else:
                            nm = month % 12 + 1
                            ny = year + (1 if month == 12 else 0)
                            end_str = f'{ny}-{nm:02d}-01 {next_cout}'
                    else:
                        end_str = f'{date_str} {c_out}' if c_out else date_str
                    duties.append({
                        'crewId': crew_id, 'seniority': seniority, 'name': name,
                        'crew_base': crew_base, 'crew_rank': crew_rank,
                        'credit': rp_cr, 'daysOff': days_off,
                        'date': date_str,
                        'startDate': f'{date_str} {c_in}' if c_in else date_str,
                        'endDate':   end_str,
                        'assignment': code if code else 'UNKNOWN',
                        'creditHours': crhour,
                        'pairingLabel': code if code else '',
                        'actingRank': act_rank,
                    })

                # ── Additional sub-blocks: same-day multi-duties ──
                # Navtech stacks extra 42pt sub-blocks (C/IN at +43, +85, +127 ...)
                # for each additional duty on the same day. Loop until no data found.
                # time_col_dict filters out column-header repeat / footer noise.
                sub_n = 2
                while sub_n <= 10:  # safety cap
                    blk_off = 42 * (sub_n - 1)
                    if block_end < block_top + blk_off + 20:
                        break
                    cin_n    = time_col_dict(get_row(blk_off + 1,  2), col_map)
                    cout_n   = time_col_dict(get_row(blk_off + 7,  2), col_map)
                    crhour_n = col_dict_split(get_row(blk_off + 13, 2), col_map)
                    code_n   = col_dict_split(get_row(blk_off + 19, 2), col_map)

                    all_days_n = set(k for k in list(code_n.keys()) + list(cin_n.keys())
                                     if isinstance(k, int))
                    if not all_days_n:
                        break

                    for day in sorted(all_days_n):
                        cn_code  = code_n.get(day, '')
                        cn_cin   = cin_n.get(day, '')
                        cn_cout  = cout_n.get(day, '')
                        cn_cr    = crhour_n.get(day, '')
                        if not cn_code and not cn_cin:
                            continue
                        date_str = f'{year}-{month:02d}-{day:02d}'
                        duties.append({
                            'crewId': crew_id, 'seniority': seniority, 'name': name,
                            'crew_base': crew_base, 'crew_rank': crew_rank,
                            'credit': rp_cr, 'daysOff': days_off,
                            'date': date_str,
                            'startDate': f'{date_str} {cn_cin}'  if cn_cin  else date_str,
                            'endDate':   f'{date_str} {cn_cout}' if cn_cout else date_str,
                            'assignment': cn_code if cn_code else 'UNKNOWN',
                            'creditHours': cn_cr,
                            'pairingLabel': cn_code if cn_code else '',
                            'actingRank': '',
                        })
                    sub_n += 1

    # Sort by seniority → date
    duties.sort(key=lambda d: (int(d['seniority']), d['date']))
    return duties


# ─── HTTP Handler ─────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

class ThreadingHTTPServer(http.server.ThreadingHTTPServer):
    pass

class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print(f'  {self.address_string()} {fmt % args}')

    def do_GET(self):
        path = urlparse(self.path).path
        if path == '/api/livereload':
            self.handle_livereload()
            return
        if path == '/api/sql-files':
            self.handle_sql_files()
            return
        if path == '/api/whitelist':
            self.handle_whitelist()
            return
        if path == '/api/analyse-status':
            self.handle_analyse_status()
            return
        if path in ('/', '/index.html'):
            path = '/ROIs_Crew_platform.html'
        filepath = os.path.join(BASE_DIR, path.lstrip('/'))
        if os.path.isfile(filepath):
            self.serve_file(filepath)
        else:
            self.send_error(404)

    def do_POST(self):
        if self.path == '/api/convert':
            self.handle_convert()
        elif self.path == '/api/analyse-sql':
            self.handle_analyse_sql()
        elif self.path == '/api/load-sql':
            self.handle_load_sql()
        elif self.path == '/api/whitelist-toggle':
            self.handle_whitelist_toggle()
        elif self.path == '/api/nbids-reformat':
            self.handle_nbids_reformat()
        else:
            self.send_error(404)

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def handle_livereload(self):
        """Server-Sent Events endpoint. Keeps connection open; sends 'reload' on file change."""
        self.send_response(200)
        self.send_header('Content-Type', 'text/event-stream')
        self.send_header('Cache-Control', 'no-cache')
        self.send_header('X-Accel-Buffering', 'no')
        self._cors()
        self.end_headers()
        q = queue.Queue(maxsize=4)
        with _lr_lock:
            _lr_clients.add(q)
        try:
            # Send an initial ping so the browser knows the connection is live
            self.wfile.write(b': connected\n\n')
            self.wfile.flush()
            while True:
                try:
                    msg = q.get(timeout=15)      # 15-s heartbeat
                    self.wfile.write(f'data: {msg}\n\n'.encode())
                except queue.Empty:
                    self.wfile.write(b': ping\n\n')  # keep-alive comment
                self.wfile.flush()
        except (BrokenPipeError, ConnectionResetError):
            pass
        finally:
            with _lr_lock:
                _lr_clients.discard(q)

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def handle_convert(self):
        try:
            content_type = self.headers.get('Content-Type', '')
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)

            all_duties = []
            crews_seen = set()

            # Multipart form: multiple PDF files
            if 'multipart/form-data' in content_type:
                boundary = content_type.split('boundary=')[-1].encode()
                parts = body.split(b'--' + boundary)
                for part in parts:
                    if b'filename=' not in part: continue
                    # Extract binary content after double CRLF
                    header_end = part.find(b'\r\n\r\n')
                    if header_end == -1: continue
                    pdf_data = part[header_end+4:].rstrip(b'\r\n--')
                    if len(pdf_data) < 100: continue
                    duties = parse_pdf_bytes(pdf_data)
                    all_duties.extend(duties)
                    for d in duties:
                        crews_seen.add(d['crewId'])
            else:
                # Raw binary body
                duties = parse_pdf_bytes(body)
                all_duties.extend(duties)
                for d in duties: crews_seen.add(d['crewId'])

            # De-duplicate (same crew+date+assignment across files)
            seen_keys = set()
            deduped   = []
            for d in all_duties:
                k = (d['crewId'], d['date'], d['assignment'])
                if k not in seen_keys:
                    seen_keys.add(k)
                    deduped.append(d)

            deduped.sort(key=lambda d: (int(d['seniority']), d['date']))

            result = {
                'duties': deduped,
                'totalDuties': len(deduped),
                'totalCrews': len(set(d['crewId'] for d in deduped)),
            }

            payload = json.dumps(result).encode()
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(payload))
            self.end_headers()
            self.wfile.write(payload)

        except Exception as e:
            err = json.dumps({'error': str(e), 'trace': traceback.format_exc()}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(err))
            self.end_headers()
            self.wfile.write(err)

    def handle_whitelist(self):
        try:
            tables = []
            if os.path.isfile(WHITELIST_PATH):
                with open(WHITELIST_PATH, 'r') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            if ':' in line:
                                name, flag = line.rsplit(':', 1)
                                load_data = flag.upper() == 'Y'
                            else:
                                name, load_data = line, False
                            tables.append({'name': name, 'load_data': load_data})
            payload = json.dumps({'tables': tables, 'count': len(tables)}).encode()
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(payload))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as e:
            err = json.dumps({'error': str(e)}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(err))
            self.end_headers()
            self.wfile.write(err)

    def handle_whitelist_toggle(self):
        """Toggle the load_data flag (Y/N) for a table in table_whitelist.md."""
        try:
            length  = int(self.headers.get('Content-Length', 0))
            body    = json.loads(self.rfile.read(length))
            table   = body.get('table', '').strip()
            enabled = bool(body.get('load_data', False))
            if not table:
                raise ValueError('table name required')

            lines = []
            if os.path.isfile(WHITELIST_PATH):
                with open(WHITELIST_PATH, 'r') as f:
                    lines = f.readlines()

            new_lines = []
            found = False
            for line in lines:
                stripped = line.rstrip('\n')
                if stripped.startswith('#') or not stripped.strip():
                    new_lines.append(line)
                    continue
                if ':' in stripped:
                    name = stripped.rsplit(':', 1)[0]
                else:
                    name = stripped
                if name == table:
                    new_lines.append(f'{name}:{"Y" if enabled else "N"}\n')
                    found = True
                else:
                    new_lines.append(line)

            if not found:
                new_lines.append(f'{table}:{"Y" if enabled else "N"}\n')

            with open(WHITELIST_PATH, 'w') as f:
                f.writelines(new_lines)

            payload = json.dumps({'ok': True, 'table': table, 'load_data': enabled}).encode()
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(payload))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as e:
            err = json.dumps({'error': str(e)}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(err))
            self.end_headers()
            self.wfile.write(err)

    def handle_sql_files(self):
        try:
            files = []
            for root, dirs, fnames in os.walk(SQL_DATA_DIR):
                dirs.sort()
                for fname in sorted(fnames):
                    full = os.path.join(root, fname)
                    rel  = os.path.relpath(full, SQL_DATA_DIR)
                    files.append({
                        'name':     fname,
                        'rel_path': rel,
                        'size':     os.path.getsize(full),
                    })
            payload = json.dumps({'files': files}).encode()
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(payload))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as e:
            err = json.dumps({'error': str(e)}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(err))
            self.end_headers()
            self.wfile.write(err)

    def handle_analyse_sql(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = json.loads(self.rfile.read(length))
            rel    = body.get('file', '')
            full   = os.path.realpath(os.path.join(SQL_DATA_DIR, rel))
            if not full.startswith(os.path.realpath(SQL_DATA_DIR) + os.sep):
                raise ValueError('Path traversal not allowed')
            if not full.endswith('.sql'):
                raise ValueError('Only .sql files are allowed')

            job_id = str(uuid.uuid4())[:8]
            _jobs[job_id] = {
                'status': 'running', 'phase': 'scanning',
                'tables': [], 'progress': 0, 'total': 0,
                'workers': ANALYSE_WORKERS, 'limit': TABLE_LIMIT, 'error': '',
                'worker_status': [{'id': i+1, 'state': 'STANDBY', 'table': ''} for i in range(ANALYSE_WORKERS)]
            }

            def _count_rows(file_path, start_off, end_off):
                count = 0
                try:
                    with open(file_path, 'rb') as fh:
                        fh.seek(start_off)
                        while True:
                            if fh.tell() >= end_off:
                                break
                            line = fh.readline()
                            if not line:
                                break
                            if line.startswith(b'INSERT INTO'):
                                count += 1
                except Exception:
                    pass
                return count

            def run():
                try:
                    whitelist = {}  # name -> load_data bool
                    if os.path.isfile(WHITELIST_PATH):
                        with open(WHITELIST_PATH) as f:
                            for line in f:
                                line = line.strip()
                                if line and not line.startswith('#'):
                                    if ':' in line:
                                        name, flag = line.rsplit(':', 1)
                                        whitelist[name] = flag.upper() == 'Y'
                                    else:
                                        whitelist[line] = False

                    # Phase 1: locate all table byte-offsets with grep -b (single fast pass)
                    _jobs[job_id]['phase'] = 'scanning'
                    for w in _jobs[job_id]['worker_status']:
                        w['state'] = 'AWAITING OFFSETS'
                    gp = subprocess.run(
                        ['grep', '-b', '^-- Table structure for ', full],
                        capture_output=True, text=True, timeout=600
                    )
                    entries = []
                    for line in gp.stdout.splitlines():
                        try:
                            colon  = line.index(':')
                            offset = int(line[:colon])
                            name   = line[colon+1:].replace('-- Table structure for ', '').strip()
                            entries.append((offset, name))
                        except Exception:
                            pass

                    # Filter to only tables in whitelist with load_data=True, then cap
                    entries = [(off, name) for off, name in entries if whitelist.get(name) is True]
                    entries = entries[:TABLE_LIMIT]
                    file_size = os.path.getsize(full)
                    _jobs[job_id]['total'] = len(entries)
                    _jobs[job_id]['phase'] = 'counting'

                    # Immediately publish all table names with sql_rows=None (pending)
                    init_tables = [
                        {'name': name, 'sql_rows': None, 'in_whitelist': name in whitelist,
                         'load_data': whitelist.get(name, False)}
                        for _, name in entries
                    ]
                    _jobs[job_id]['tables'] = init_tables

                    # Phase 2: parallel row count — workers update counts in-place as they finish
                    tasks = []
                    for i, (offset, name) in enumerate(entries):
                        end_off = entries[i+1][0] if i+1 < len(entries) else file_size
                        tasks.append((full, offset, end_off, name, i))

                    lock = threading.Lock()
                    # Map thread ident → worker slot (0-based)
                    thread_slot = {}
                    slot_counter = [0]

                    for w in _jobs[job_id]['worker_status']:
                        w['state'] = 'IDLE'
                        w['table'] = ''

                    def worker(task):
                        file_path, start_off, end_off, name, idx = task
                        tid = threading.current_thread().ident

                        # Assign a stable worker slot to this thread
                        with lock:
                            if tid not in thread_slot:
                                thread_slot[tid] = slot_counter[0] % ANALYSE_WORKERS
                                slot_counter[0] += 1
                            slot = thread_slot[tid]
                            ws = _jobs[job_id]['worker_status'][slot]
                            ws['state'] = 'COUNTING'
                            ws['table'] = name

                        count = _count_rows(file_path, start_off, end_off)

                        with lock:
                            _jobs[job_id]['worker_status'][slot]['state'] = 'DONE'
                            _jobs[job_id]['worker_status'][slot]['table'] = name
                            _jobs[job_id]['tables'][idx]['sql_rows'] = count
                            _jobs[job_id]['progress'] = sum(
                                1 for t in _jobs[job_id]['tables'] if t['sql_rows'] is not None
                            )

                    with concurrent.futures.ThreadPoolExecutor(max_workers=ANALYSE_WORKERS) as ex:
                        list(ex.map(worker, tasks))

                    for w in _jobs[job_id]['worker_status']:
                        w['state'] = 'DONE'

                    _jobs[job_id]['progress'] = len(entries)
                    _jobs[job_id]['status']   = 'done'

                except Exception as ex:
                    _jobs[job_id]['status'] = 'error'
                    _jobs[job_id]['error']  = str(ex)

            threading.Thread(target=run, daemon=True).start()

            payload = json.dumps({'job_id': job_id}).encode()
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(payload))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as e:
            err = json.dumps({'error': str(e), 'trace': traceback.format_exc()}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(err))
            self.end_headers()
            self.wfile.write(err)

    def handle_analyse_status(self):
        try:
            qs     = parse_qs(urlparse(self.path).query)
            job_id = qs.get('job', [''])[0]
            job    = _jobs.get(job_id)
            if not job:
                raise ValueError(f'Unknown job: {job_id}')
            payload = json.dumps(job).encode()
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(payload))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as e:
            err = json.dumps({'error': str(e)}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(err))
            self.end_headers()
            self.wfile.write(err)

    def handle_load_sql(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = json.loads(self.rfile.read(length))
            rel    = body.get('file', '')
            full   = os.path.realpath(os.path.join(SQL_DATA_DIR, rel))
            if not full.startswith(os.path.realpath(SQL_DATA_DIR) + os.sep):
                raise ValueError('Path traversal not allowed')
            if not full.endswith('.sql'):
                raise ValueError('Only .sql files are allowed')

            split_script  = os.path.join(BASE_DIR, 'split_sql.py')
            import_script = os.path.join(SQL_DATA_DIR, 'run_import.sh')
            tables_dir    = os.path.join(SQL_DATA_DIR, 'tables')

            # Step 1: split in background
            subprocess.Popen(
                ['bash', '-c',
                 f'python3 {split_script} {full} --out-dir {tables_dir} --quiet'
                 f' && bash {import_script}'
                 f' >> /tmp/royce-loadsql.log 2>&1'],
                close_fds=True
            )
            result = type('R', (), {'returncode': 0, 'stdout': 'Split + import started in background. Check /tmp/royce-loadsql.log', 'stderr': ''})()
            payload = json.dumps({
                'exit_code': result.returncode,
                'stdout':    result.stdout,
                'stderr':    result.stderr,
            }).encode()
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(payload))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as e:
            err = json.dumps({'error': str(e), 'trace': traceback.format_exc()}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(err))
            self.end_headers()
            self.wfile.write(err)

    def handle_nbids_reformat(self):
        try:
            content_type = self.headers.get('Content-Type', '')
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)

            txt_files   = []   # list of (filename, raw_bytes)
            period_override = ''

            if 'multipart/form-data' in content_type:
                boundary = content_type.split('boundary=')[-1].encode()
                parts = body.split(b'--' + boundary)
                for part in parts:
                    if b'name="file"' in part and b'filename=' in part:
                        # Extract original filename from Content-Disposition
                        fname_m = re.search(rb'filename="([^"]+)"', part)
                        orig_name = fname_m.group(1).decode('utf-8', errors='replace') if fname_m else 'file.txt'
                        hend = part.find(b'\r\n\r\n')
                        if hend != -1:
                            txt_files.append((orig_name, part[hend+4:].rstrip(b'\r\n--')))
                    elif b'name="period"' in part:
                        hend = part.find(b'\r\n\r\n')
                        if hend != -1:
                            period_override = part[hend+4:].rstrip(b'\r\n--').decode('utf-8', errors='replace').strip()

            if not txt_files:
                raise ValueError('No file content received')

            # Parse each file; offset group_ids so they stay globally unique
            all_output_rows, all_error_rows = [], []
            total_input_rows = 0
            group_id_offset  = 0

            for orig_name, raw in txt_files:
                txt_content = raw.decode('utf-8', errors='replace')
                period = period_override if period_override else _nb_detect_period(txt_content)
                out_rows, err_rows, stats = nb_parse_txt(txt_content, period)
                # Apply offset to group_id for global uniqueness across files
                for r in out_rows:
                    r = dict(r)
                    r['group_id'] = r['group_id'] + group_id_offset
                    all_output_rows.append(r)
                all_error_rows.extend(err_rows)
                total_input_rows += stats['input_rows']
                group_id_offset  += stats['groups']

            # Sort rows: earlier period first, preserve original crew order within each period
            _MONTH_ORDER = {m: i for i, m in enumerate(
                ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'])}
            def _period_sort_key(r):
                p = (r.get('period') or '').strip()
                parts = p.split()
                mon = _MONTH_ORDER.get(parts[0][:3].capitalize(), 99) if parts else 99
                yr  = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 9999
                return (yr, mon)

            # Stable sort: preserves relative order of rows with the same period key
            all_output_rows.sort(key=_period_sort_key)

            # Use combined period label (unique periods joined, chronological)
            periods_seen = list(dict.fromkeys(r['period'] for r in all_output_rows))
            combined_period = ', '.join(periods_seen)

            xlsx_bytes = _nb_build_excel(all_output_rows, all_error_rows, combined_period)
            xlsx_b64   = base64.b64encode(xlsx_bytes).decode('ascii')

            from datetime import datetime as _dt
            ts       = _dt.now().strftime('%Y-%m-%d-%H%M%S')
            filename = f'crew_bids_reference-{ts}.xlsx'

            payload = json.dumps({
                'xlsx_b64':    xlsx_b64,
                'filename':    filename,
                'input_rows':  total_input_rows,
                'parsed_rows': len(all_output_rows),
                'groups':      group_id_offset,
                'error_rows':  len(all_error_rows),
                'files':       len(txt_files),
            }).encode()

            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(payload))
            self.end_headers()
            self.wfile.write(payload)

        except Exception as e:
            err = json.dumps({'error': str(e), 'trace': traceback.format_exc()}).encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(err))
            self.end_headers()
            self.wfile.write(err)

    def serve_file(self, path):
        ext = os.path.splitext(path)[1].lower()
        mime = {'.html':'text/html','.js':'application/javascript',
                '.css':'text/css','.json':'application/json',
                '.svg':'image/svg+xml','.png':'image/png',
                '.jpg':'image/jpeg','.jpeg':'image/jpeg'}.get(ext,'application/octet-stream')
        with open(path, 'rb') as f:
            data = f.read()
        self.send_response(200)
        self.send_header('Content-Type', mime)
        self.send_header('Content-Length', len(data))
        self.end_headers()
        self.wfile.write(data)


# ─── N-Bids Parsing Logic ─────────────────────────────────────────────────────

_MONTH_ABBREV = {
    'january':'Jan','february':'Feb','march':'Mar','april':'Apr',
    'may':'May','june':'Jun','july':'Jul','august':'Aug',
    'september':'Sep','october':'Oct','november':'Nov','december':'Dec',
}

def _nb_detect_period(txt):
    """Extract and normalise period from 'Period: <value>' in the TXT header."""
    for line in txt.splitlines()[:20]:
        m = re.match(r'^\s*Period\s*:\s*(.+)', line, re.I)
        if m:
            raw = m.group(1).strip()
            # Normalise full month name → 3-letter abbrev (e.g. "December 2025" → "Dec 2025")
            parts = raw.split()
            if parts:
                abbrev = _MONTH_ABBREV.get(parts[0].lower())
                if abbrev:
                    parts[0] = abbrev
                    return ' '.join(parts)
            return raw
    return 'Unknown'

_AE = r'(?:Any|Every)'
_DATE_RE = re.compile(
    r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s+\d{4}', re.I)

_PAIRING_PROPS = [
    ( 3, re.compile(r'^Pairing Check-In Time',                  re.I)),
    (13, re.compile(r'^Pairing Check-Out Time',                 re.I)),
    (16, re.compile(r'^Pairing Length',                         re.I)),
    ( 2, re.compile(r'^Pairing Number',                         re.I)),
    (32, re.compile(r'^Departing On',                           re.I)),
    (45, re.compile(r'^' + _AE + r'\s+Duty Legs\s+Counting Deadhead', re.I)),
    ( 8, re.compile(r'^' + _AE + r'\s+Duty Legs',              re.I)),
    ( 1, re.compile(r'^' + _AE + r'\s+Landing In',             re.I)),
    (44, re.compile(r'^' + _AE + r'\s+Duty In',                re.I)),
    (19, re.compile(r'^' + _AE + r'\s+Layover In',             re.I)),
    (30, re.compile(r'^' + _AE + r'\s+Duty Duration',          re.I)),
    ( 6, re.compile(r'^TAFB',                                   re.I)),
    ( 9, re.compile(r'^Average Daily Credit',                   re.I)),
    (46, re.compile(r'^Average Credit',                         re.I)),
    ( 4, re.compile(r'^(?:Pairing\s+)?Total Credit',            re.I)),
    (31, re.compile(r'^' + _AE + r'\s+Duty On Time',           re.I)),
    (37, re.compile(r'^' + _AE + r'\s+Duty On(?:\s+Date)?',   re.I)),
    (28, re.compile(r'^Total Legs In First Duty',               re.I)),
    (42, re.compile(r'^Total Legs In Last Duty',                re.I)),
    (11, re.compile(r'^Total Legs In Pairing',                  re.I)),
    (26, re.compile(r'^' + _AE + r'\s+Flight Number',          re.I)),
    (25, re.compile(r'^' + _AE + r'\s+Leg Is Redeye',          re.I)),
    (34, re.compile(r'^Credit Per Time Away From Base',         re.I)),
    (23, re.compile(r'^' + _AE + r'\s+Enroute Check-In Time',  re.I)),
    (38, re.compile(r'^' + _AE + r'\s+Enroute Check-Out Time', re.I)),
    (22, re.compile(r'^' + _AE + r'\s+Layover Of Duration',    re.I)),
    (49, re.compile(r'^' + _AE + r'\s+Layover On(?:\s+Date)?', re.I)),
    (27, re.compile(r'^' + _AE + r'\s+Leg With Employee Number', re.I)),
    (29, re.compile(r'^Deadhead Legs',                          re.I)),
    (33, re.compile(r'^Average Daily Block Time',               re.I)),
    (36, re.compile(r'^(?:Pairing\s+)?Total Block Time',        re.I)),
    (40, re.compile(r'^Deadhead Day',                           re.I)),
    (41, re.compile(r'^' + _AE + r'\s+Sit Length',             re.I)),
]

_SKIP_BIDS = {'Pairing Bid Group', 'Reserve Bid Group'}

def _nb_norm_val(v):
    v = str(v).strip()
    v = re.sub(r'\s+(legs?|days?)\s*$', '', v, flags=re.I).strip()
    return v or None

def _nb_norm_list(v):
    parts = [p.strip() for p in str(v).split(',')]
    return ','.join(p for p in parts if p) or None

def _nb_strip_modifiers(s):
    limit_n = all_or_nothing = None
    s = re.sub(r'\s*Else Start Next Bid Group\s*$', '', s, flags=re.I).strip()
    m = re.search(r'\s*Limit\s+(\d+)\s*$', s, re.I)
    if m:
        limit_n = int(m.group(1))
        s = s[:m.start()].strip()
    if re.search(r'\bAll or Nothing\b', s, re.I):
        all_or_nothing = 1
        s = re.sub(r'\s*All or Nothing\s*', ' ', s, flags=re.I).strip()
    return s, limit_n, all_or_nothing

def _nb_extract_op_params(rem):
    s = rem.strip()
    m = re.match(r'^Between\s+(.+?)\s+And\s+(.+)$', s, re.I)
    if m:
        return 'Between', _nb_norm_val(m.group(1)), _nb_norm_val(m.group(2)), None
    m = re.match(r'^(>=|<=|>|<|=)\s+(.+)$', s)
    if m:
        return m.group(1), _nb_norm_val(m.group(2)), None, None
    if s:
        return None, _nb_norm_list(s), None, None
    return None, None, None, None

def _nb_match_pairing_prop(clause):
    clause = clause.strip()
    for pid, pat in _PAIRING_PROPS:
        m = pat.match(clause)
        if m:
            return pid, clause[m.end():].strip()
    return None, None

def _nb_parse_prefer_off(s):
    body = re.sub(r'^Prefer Off\s*', '', s, flags=re.I).strip()
    minimum_n = None
    m = re.search(r'\s+Minimum\s+(\d+)\s*$', body, re.I)
    if m:
        minimum_n = int(m.group(1))
        body = body[:m.start()].strip()
    param_b = param_c = None
    tw = re.search(r'\s+Between\s+(\d{1,3}:\d{2})\s+And\s+(\d{1,3}:\d{2})\s*$', body, re.I)
    if tw:
        param_b, param_c = tw.group(1), tw.group(2)
        body = body[:tw.start()].strip()
    if re.match(r'^Between\b', body, re.I):
        m = re.match(r'^Between\s+(.+?)\s+And\s+(.+)$', body, re.I)
        if m:
            return 'Between', m.group(1).strip(), m.group(2).strip(), param_c, minimum_n
    dp = r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s+\d{4}'
    dr = re.match(r'^(' + dp + r')\s*[-\u2013]\s*(' + dp + r')\s*$', body, re.I)
    if dr:
        return 'Between', dr.group(1).strip(), dr.group(2).strip(), param_c, minimum_n
    th = re.match(r'^(.+?)\s+Through\s+(.+)$', body, re.I)
    if th:
        return 'Between', th.group(1).strip(), th.group(2).strip(), param_c, minimum_n
    dates = _DATE_RE.findall(body)
    if dates:
        return None, ','.join(d.strip() for d in dates), param_b, param_c, minimum_n
    return None, _nb_norm_list(body) if body else None, param_b, param_c, minimum_n

def _nb_parse_set_condition(s):
    m = re.match(
        r'^Set Condition\s+(\d+)\s+Consecutive Days Off In A Row'
        r'(?:\s+Between\s+(.+?)\s+And\s+(.+))?$', s, re.I)
    if m:
        return 12, ('Between' if m.group(2) else None), m.group(1), m.group(2), m.group(3)
    m = re.match(r'^Set Condition Minimum Days Off In A Row\s+(\d+)$', s, re.I)
    if m: return 14, None, m.group(1), None, None
    m = re.match(r'^Set Condition Maximum Days Off In A Row\s+(\d+)$', s, re.I)
    if m: return 48, None, m.group(1), None, None
    m = re.match(r'^Set Condition Maximum Days On In A Row\s+(\d+)$', s, re.I)
    if m: return 15, None, m.group(1), None, None
    m = re.match(
        r'^Set Condition Pattern Between\s+(\d+)\s+and\s+(\d+)\s+Days On'
        r',?\s+with\s+(\d+)\s+Days Off', s, re.I)
    if m: return 21, 'Between', m.group(3), m.group(1), m.group(2)
    m = re.match(r'^Set Condition Days Off Opposite Employee\s+(\d+)\s+Minimum\s+(\d+)$', s, re.I)
    if m: return 43, None, m.group(1), m.group(2), None
    if re.match(r'^Set Condition Minimum Credit Window$', s, re.I): return 18, None, None, None, None
    if re.match(r'^Set Condition Maximum Credit Window$', s, re.I): return 17, None, None, None, None
    m = re.match(r'^Set Condition Minimum Base Layover\s+(\S+)$', s, re.I)
    if m: return 39, None, m.group(1), None, None
    if re.match(r'^Set Condition No Same Day Pairings$', s, re.I): return 20, None, None, None, None
    m = re.match(r'^Set Condition Short Call Type\s+(\S+)$', s, re.I)
    if m: return 5, None, m.group(1), None, None
    return None, None, None, None, None

def _nb_parse_row(raw):
    """Returns list of node dicts, [] to skip, or None on error."""
    s, ln, aon = _nb_strip_modifiers(raw)
    if s in _SKIP_BIDS:
        return []
    if re.match(r'^Clear Schedule and Start Next Bid Group$', s, re.I):
        return [dict(action_id=None, property_id=10, operator=None, param_a=None, param_b=None, param_c=None,
                     node_id=1, and_or_or=None, limit_n=ln, all_or_nothing=aon, minimum_n=None)]
    m = re.match(r'^Forget Line\s+(\d+)$', s, re.I)
    if m:
        return [dict(action_id=None, property_id=35, operator=None, param_a=m.group(1), param_b=None, param_c=None,
                     node_id=1, and_or_or=None, limit_n=ln, all_or_nothing=aon, minimum_n=None)]
    if re.match(r'^Waive No Same Day Duty Starts$', s, re.I):
        return [dict(action_id=None, property_id=24, operator=None, param_a=None, param_b=None, param_c=None,
                     node_id=1, and_or_or=None, limit_n=ln, all_or_nothing=aon, minimum_n=None)]
    if re.match(r'^Prefer Off\b', s, re.I):
        op, pa, pb, pc, mn = _nb_parse_prefer_off(s)
        return [dict(action_id=None, property_id=7, operator=op, param_a=pa, param_b=pb, param_c=pc,
                     node_id=1, and_or_or=None, limit_n=ln, all_or_nothing=aon, minimum_n=mn)]
    m = re.match(r'^Award Reserve Day On\s+(.+)$', s, re.I)
    if m:
        dates = _DATE_RE.findall(m.group(1))
        pa = ','.join(d.strip() for d in dates) if dates else _nb_norm_list(m.group(1))
        return [dict(action_id=None, property_id=47, operator=None, param_a=pa, param_b=None, param_c=None,
                     node_id=1, and_or_or=None, limit_n=ln, all_or_nothing=aon, minimum_n=None)]
    if re.match(r'^Set Condition\b', s, re.I):
        pid, op, pa, pb, pc = _nb_parse_set_condition(s)
        if pid:
            return [dict(action_id=None, property_id=pid, operator=op, param_a=pa, param_b=pb, param_c=pc,
                         node_id=1, and_or_or=None, limit_n=ln, all_or_nothing=aon, minimum_n=None)]
        return None
    ma = re.match(r'^Award Pairings\s+If\s+', s, re.I)
    mv = re.match(r'^Avoid Pairings\s+If\s+', s, re.I)
    if ma or mv:
        ai   = 1 if ma else 2
        rest = s[(ma or mv).end():]
        clauses = re.split(r'\s+If\s+', rest, flags=re.I)
        nodes = []
        for i, clause in enumerate(clauses):
            clause = clause.strip()
            if not clause: continue
            if re.match(r'^Followed By Pairings', clause, re.I): break
            pid, rem = _nb_match_pairing_prop(clause)
            if pid is None: return None
            op, pa, pb, pc = _nb_extract_op_params(rem)
            nodes.append(dict(
                action_id=(ai if i == 0 else None),
                property_id=pid, operator=op, param_a=pa, param_b=pb, param_c=pc,
                node_id=i + 1, and_or_or=(None if i == 0 else 'AND'),
                limit_n=(ln if i == 0 else None),
                all_or_nothing=(aon if i == 0 else None),
                minimum_n=None))
        return nodes if nodes else None
    return None

def _nb_build_excel(rows, errors, period):
    """Build Excel workbook bytes using openpyxl.
    Sheet 'crew_bids' matches the reference layout exactly:
      Row 1 = column headers, Row 2 = description, Row 3+ = data
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    # Sheet 1: crew_bids  (matches reference column layout)
    ws1 = wb.active
    ws1.title = 'crew_bids'
    headers1 = ['id','crew_id','bid_context','period','layer','property_group_id',
                 'node_id','and_or_or','action_id','property_id','operator',
                 'param_a','param_b','param_c','limit_n','all_or_nothing','minimum_n']
    desc_row  = ['BIGINT AUTO_INCREMENT','MEDIUMINT UNSIGNED','Default | Current',
                 'e.g. Dec 2025','1 \u2026 24',None,
                 '1 = action row\n2+ = chained condition',
                 'AND | OR\n(NULL for node_id=1)',
                 '1=Award  2=Avoid\nNULL for non-Pairing',
                 '\u2192 bid_properties.id',
                 '> < >= <= =\nBetween | NULL',
                 'value or lower bound\n(list: comma-separated)',
                 'upper bound\n(Between only)',
                 '3rd value\n(if needed)',
                 'Limit N\n(max pairings to award)',
                 '1 = All or Nothing\nmodifier on Prefer Off',
                 'Minimum N\nfor Prefer Off Weekends']
    # Row 1 — header: Calibri 11pt bold white on #1F4E79, center+center+wrap
    hdr_font = Font(bold=True, color='00FFFFFF', size=11, name='Calibri')
    hdr_fill = PatternFill('solid', fgColor='001F4E79')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    # Row 2 — description: Calibri 10pt grey (#595959) on #FFF2CC, left+top+wrap
    desc_font  = Font(bold=False, color='00595959', size=10, name='Calibri')
    desc_fill  = PatternFill('solid', fgColor='00FFF2CC')
    desc_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws1.append(desc_row)
    for cell in ws1[2]:
        cell.font  = desc_font
        cell.fill  = desc_fill
        cell.alignment = desc_align

    # Internal key → header mapping
    _col_map = {
        'id':                 None,           # always None
        'crew_id':            'crew_id',
        'bid_context':        'bid_context',
        'period':             'period',
        'layer':              'layer',
        'property_group_id':  'group_id',
        'node_id':            'node_id',
        'and_or_or':          'and_or',
        'action_id':          'action_id',
        'property_id':        'property_id',
        'operator':           'operator',
        'param_a':            'param_a',
        'param_b':            'param_b',
        'param_c':            'param_c',
        'limit_n':            'limit_n',
        'all_or_nothing':     'all_or_nothing',
        'minimum_n':          'minimum_n',
    }
    # Data rows — Calibri 10pt, no fill, left+top alignment
    data_font  = Font(bold=False, size=10, name='Calibri')
    data_align = Alignment(horizontal='left', vertical='top')
    for r in rows:
        row_out = []
        for h in headers1:
            key = _col_map[h]
            if key is None:
                row_out.append(None)
            elif h == 'crew_id':
                row_out.append(int(r.get(key)) if r.get(key) is not None else None)
            else:
                row_out.append(r.get(key))
        ws1.append(row_out)
    for row in ws1.iter_rows(min_row=3):
        for cell in row:
            cell.font      = data_font
            cell.alignment = data_align

    # Exact column widths matching reference
    _col_widths = {'A':8,'B':10,'C':14,'D':14,'E':8,'F':18,
                   'G':10,'H':10,'I':12,'J':13,'K':14,
                   'L':16,'M':16,'N':12,'O':10,'P':15,'Q':12}
    for col_letter, width in _col_widths.items():
        ws1.column_dimensions[col_letter].width = width

    # Freeze panes at A3 (matching reference)
    ws1.freeze_panes = 'A3'

    # Sheet 2: Errors
    ws2 = wb.create_sheet('Errors')
    headers2 = ['crew_id','bid_context','layer','raw_bid']
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = hdr_font
        cell.fill = PatternFill('solid', fgColor='5E2020')
        cell.alignment = Alignment(horizontal='center')
    for e in errors:
        ws2.append([e.get(h) for h in headers2])
    for col in ws2.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def nb_parse_txt(txt_content, period):
    """Parse PBS bid report TXT format. Returns (output_rows, error_rows, stats).

    Expected format:
      ------------...        <- top-level separator (no leading spaces, 10+ dashes)
      Seniority N  Category ...  Employee # NNN
      Confirmation: ...  [Default|Current] Bid
      ------------...
      Buddies:
      Bid Preferences:
             ---...          <- block separator (leading spaces, 3+ dashes)
         1.  <bid string>
         ...
             ---...          <- next block (layer 2)
         N.  <bid string>
    """
    lines       = txt_content.splitlines()
    output_rows = []
    error_rows  = []
    group_id    = 0
    input_rows  = 0
    n_lines     = len(lines)

    CREW_DASH_RE   = re.compile(r'^-{10,}\s*$')            # top-level separator
    BLOCK_DASH_RE  = re.compile(r'^\s+-{3,}\s*$')          # indented block separator
    EMP_RE         = re.compile(r'Employee\s*#\s+(\d+)', re.I)
    BIDTYPE_RE     = re.compile(r'\b(Default|Current)\s+Bid\b', re.I)
    ITEM_RE        = re.compile(r'^\s{1,10}(\d+)\.\s+(.+)$')
    SKIP_MARKER_RE = re.compile(r'^(Award|Avoid)\s+Pairings\s*$', re.I)

    def emit(crew_id, bid_context, layer, bid_str):
        nonlocal group_id, input_rows
        if not bid_str or bid_str in _SKIP_BIDS:
            return
        if SKIP_MARKER_RE.match(bid_str):
            return
        input_rows += 1
        nodes = _nb_parse_row(bid_str)
        if nodes is None:
            error_rows.append({'crew_id': crew_id, 'bid_context': bid_context,
                                'layer': layer, 'raw_bid': bid_str})
            return
        if not nodes:
            return
        group_id += 1
        for nd in nodes:
            output_rows.append({
                'crew_id':        crew_id,
                'bid_context':    bid_context,
                'period':         period,
                'layer':          layer,
                'group_id':       group_id,
                'node_id':        nd['node_id'],
                'and_or':         nd['and_or_or'],
                'action_id':      nd['action_id'],
                'property_id':    nd['property_id'],
                'operator':       nd['operator'],
                'param_a':        nd['param_a'],
                'param_b':        nd['param_b'],
                'param_c':        nd['param_c'],
                'limit_n':        nd['limit_n'],
                'all_or_nothing': nd['all_or_nothing'],
                'minimum_n':      nd['minimum_n'],
                'raw_bid':        bid_str,
            })

    i = 0
    while i < n_lines:
        raw = lines[i].rstrip()

        if not CREW_DASH_RE.match(raw):
            i += 1
            continue

        # Top-level separator found — next non-empty line should be crew header
        i += 1
        while i < n_lines and not lines[i].strip():
            i += 1
        if i >= n_lines:
            break

        emp_m = EMP_RE.search(lines[i])
        if not emp_m:
            # Not a crew header (e.g., end-of-file separator)
            continue

        crew_id     = str(int(emp_m.group(1)))  # strip leading zeros
        bid_context = 'Default'
        if i + 1 < n_lines:
            bm = BIDTYPE_RE.search(lines[i + 1])
            if bm:
                bid_context = bm.group(1).capitalize()

        # Advance past header + confirmation line, then scan for "Bid Preferences:"
        i += 2
        while i < n_lines:
            if CREW_DASH_RE.match(lines[i].rstrip()):
                # Another top-level dash line (end-of-header dash) — skip it
                i += 1
                continue
            if 'Bid Preferences:' in lines[i]:
                i += 1  # move past "Bid Preferences:"
                break
            i += 1
        else:
            continue  # no Bid Preferences found

        # Parse bid blocks — each block starts with a BLOCK_DASH_RE line
        layer = 0
        while i < n_lines:
            raw = lines[i].rstrip()

            if CREW_DASH_RE.match(raw):
                # End of this crew section
                break

            if BLOCK_DASH_RE.match(raw):
                layer += 1
                i += 1
                continue

            if layer > 0:
                item_m = ITEM_RE.match(raw)
                if item_m:
                    emit(crew_id, bid_context, layer, item_m.group(2).strip())

            i += 1

    return output_rows, error_rows, {'input_rows': input_rows, 'groups': group_id}

if __name__ == '__main__':
    PORT = 8088
    server = ThreadingHTTPServer(('0.0.0.0', PORT), Handler)
    print(f'NAVOPS backend running → http://localhost:{PORT}/')
    print(f'API endpoint: POST http://localhost:{PORT}/api/convert')
    print('Press Ctrl+C to stop.')
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\nStopped.')
