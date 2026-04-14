#!/usr/bin/env python3
"""
N-Bids parser test: compare output against crew_bids_reference.xlsx
Run: python3 scripts/test_nbids.py
"""
import sys, re, os, datetime
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ── Import parsing functions from server.py ────────────────────────────────
# We import selectively to avoid pdfplumber dependency
import importlib.util, types

def _load_nb_functions():
    """Extract just the _nb_* and nb_parse_txt functions from server.py."""
    src_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'server.py')
    with open(src_path, 'r') as f:
        src = f.read()
    # Create a minimal module with only what we need
    mod = types.ModuleType('nb_funcs')
    mod.__dict__['re'] = re
    mod.__dict__['io'] = __import__('io')
    # Execute only the nb-related code blocks
    # Find start of nb parsing section
    start = src.find('# ─── N-Bids Parsing Logic')
    end   = src.find('\nif __name__')
    if start == -1:
        raise RuntimeError('Cannot find N-Bids Parsing Logic section in server.py')
    nb_src = src[start:end]
    exec(compile(nb_src, 'server.py', 'exec'), mod.__dict__)
    return mod

nb = _load_nb_functions()
nb_parse_txt = nb.nb_parse_txt

# ── Load reference ─────────────────────────────────────────────────────────
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TXT_FILE = os.path.join(BASE, 'Design', 'New Crew BIds', 'Dec Crew Bids 2025 All in one.txt')
REF_FILE = os.path.join(BASE, 'Design', 'New Crew BIds', 'crew_bids_reference.xlsx')

def load_reference():
    wb  = openpyxl.load_workbook(REF_FILE, read_only=True, data_only=True)
    ws  = wb['crew_bids']
    all_rows = list(ws.iter_rows(values_only=True))
    # Row 0: column headers, Row 1: description row — skip both
    # Columns: id, crew_id, bid_context, period, layer, property_group_id,
    #           node_id, and_or_or, action_id, property_id, operator,
    #           param_a, param_b, param_c, limit_n, all_or_nothing, minimum_n
    COL = {h: i for i, h in enumerate(all_rows[0])}
    data = []
    for r in all_rows[2:]:
        if r[COL['crew_id']] is None:
            continue
        data.append({
            'crew_id':       str(r[COL['crew_id']]),
            'bid_context':   r[COL['bid_context']],
            'period':        r[COL['period']],
            'layer':         r[COL['layer']],
            'group_id':      r[COL['property_group_id']],
            'node_id':       r[COL['node_id']],
            'and_or':        r[COL['and_or_or']],
            'action_id':     r[COL['action_id']],
            'property_id':   r[COL['property_id']],
            'operator':      r[COL['operator']],
            'param_a':       r[COL['param_a']],
            'param_b':       r[COL['param_b']],
            'param_c':       r[COL['param_c']],
            'limit_n':       r[COL['limit_n']],
            'all_or_nothing':r[COL['all_or_nothing']],
            'minimum_n':     r[COL['minimum_n']],
        })
    return data

def load_txt():
    with open(TXT_FILE, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    rows, errors, stats = nb_parse_txt(content, 'Dec 2025')
    return rows, errors, stats

def normalise(v):
    """Normalise a field value for comparison."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v

CMP_FIELDS = ['crew_id','bid_context','layer','node_id','and_or',
              'action_id','property_id','operator',
              'param_a','param_b','param_c',
              'limit_n','all_or_nothing','minimum_n']

def rows_key(r):
    return (str(r['crew_id']), str(r.get('bid_context','')),
            int(r['layer']) if r['layer'] else 0,
            int(r['group_id']) if r['group_id'] else 0,
            int(r['node_id']) if r['node_id'] else 0)

def cell_str(v):
    """Normalise a cell value for comparison: None stays None, strings stripped."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def main():
    print('Parsing TXT...')
    out_rows, error_rows, stats = load_txt()
    print(f'  Output rows: {len(out_rows)}')
    print(f'  Parse errors: {len(error_rows)}')
    print(f'  Input rows processed: {stats["input_rows"]}')
    print(f'  Groups created: {stats["groups"]}')

    # Build Excel from parsed rows
    _nb_build_excel = nb._nb_build_excel
    xlsx_bytes = _nb_build_excel(out_rows, error_rows, 'Dec 2025')

    # Load both workbooks for cell-by-cell comparison
    import openpyxl, io as _io
    wb_gen = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    wb_ref = openpyxl.load_workbook(REF_FILE, read_only=True, data_only=True)

    gen_rows = list(wb_gen['crew_bids'].iter_rows(values_only=True))
    ref_rows = list(wb_ref['crew_bids'].iter_rows(values_only=True))

    print(f'\nComparing crew_bids sheet cell-by-cell...')
    print(f'  Reference: {len(ref_rows)} rows x {len(ref_rows[0])} cols')
    print(f'  Generated: {len(gen_rows)} rows x {len(gen_rows[0])} cols')

    mismatches = []
    max_rows = max(len(ref_rows), len(gen_rows))

    for ri in range(max_rows):
        if ri >= len(ref_rows):
            mismatches.append((ri+1, None, None, 'EXTRA ROW in generated'))
            continue
        if ri >= len(gen_rows):
            mismatches.append((ri+1, None, None, 'MISSING ROW in generated'))
            continue
        ref_r = ref_rows[ri]
        gen_r = gen_rows[ri]
        max_cols = max(len(ref_r), len(gen_r))
        for ci in range(max_cols):
            rv = cell_str(ref_r[ci] if ci < len(ref_r) else None)
            gv = cell_str(gen_r[ci] if ci < len(gen_r) else None)
            if rv != gv:
                col_letter = chr(ord('A') + ci) if ci < 26 else f'C{ci}'
                mismatches.append((ri+1, col_letter, rv, gv))

    print(f'\n── RESULTS ──────────────────────────────────────────')
    print(f'  Cell mismatches: {len(mismatches)}')

    if mismatches:
        print(f'\n── CELL VALUE MISMATCHES (first 60) ──')
        for row_n, col, rv, gv in mismatches[:60]:
            if col is None:
                print(f'  Row {row_n}: {rv or gv}')
            else:
                print(f'  [{col}{row_n}]  ref={repr(rv)}  gen={repr(gv)}')
        return False

    # ── Formatting comparison ────────────────────────────────────────────────
    print('\nComparing formatting...')
    import openpyxl as _opx
    wb_gen2 = _opx.load_workbook(_io.BytesIO(xlsx_bytes), data_only=True)
    wb_ref2 = _opx.load_workbook(REF_FILE, data_only=True)
    ws_gen2 = wb_gen2['crew_bids']
    ws_ref2 = wb_ref2['crew_bids']

    fmt_mismatches = []

    def _fmt(cell):
        f   = cell.font
        fi  = cell.fill
        al  = cell.alignment
        fc  = (f.color.rgb if f.color and f.color.type == 'rgb' else None) if f.color else None
        fg  = (fi.fgColor.rgb if fi.fgColor and fi.fgColor.type == 'rgb' else None) if fi and fi.fgColor else None
        return {
            'bold':    f.bold, 'color': fc, 'size': f.size, 'name': f.name,
            'fill':    fi.fill_type if fi else None, 'fgColor': fg,
            'h_align': al.horizontal, 'v_align': al.vertical, 'wrap': al.wrap_text,
        }

    # Check rows 1-3 only for formatting (header, desc, first data row)
    for ri in [1, 2, 3]:
        ref_row = list(ws_ref2.iter_rows(min_row=ri, max_row=ri))[0]
        gen_row = list(ws_gen2.iter_rows(min_row=ri, max_row=ri))[0]
        for ref_cell, gen_cell in zip(ref_row, gen_row):
            rf = _fmt(ref_cell)
            gf = _fmt(gen_cell)
            for attr in rf:
                if rf[attr] != gf[attr]:
                    fmt_mismatches.append(
                        f'  [{ref_cell.column_letter}{ri}] {attr}: ref={repr(rf[attr])} gen={repr(gf[attr])}'
                    )

    # Column widths
    ref_dims = ws_ref2.column_dimensions
    gen_dims = ws_gen2.column_dimensions
    for col in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q']:
        rw = ref_dims[col].width if col in ref_dims else None
        gw = gen_dims[col].width if col in gen_dims else None
        if rw != gw:
            fmt_mismatches.append(f'  Col {col} width: ref={rw} gen={gw}')

    # Freeze panes
    if ws_ref2.freeze_panes != ws_gen2.freeze_panes:
        fmt_mismatches.append(f'  freeze_panes: ref={ws_ref2.freeze_panes} gen={ws_gen2.freeze_panes}')

    print(f'  Formatting mismatches: {len(fmt_mismatches)}')

    if fmt_mismatches:
        print(f'\n── FORMATTING MISMATCHES ──')
        for m in fmt_mismatches:
            print(m)
        return False

    print('\n  PERFECT MATCH -- 100%  (values + formatting)')
    ts = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
    out_path = os.path.join(BASE, 'Design', 'New Crew BIds', f'crew_bids_reference-{ts}.xlsx')
    with open(out_path, 'wb') as f:
        f.write(xlsx_bytes)
    print(f'  Saved: {out_path}')
    return True

if __name__ == '__main__':
    ok = main()
    sys.exit(0 if ok else 1)
