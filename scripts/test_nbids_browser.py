#!/usr/bin/env python3
"""
Browser-simulation test for N-Bids Reformat.

Simulates the user's browser workflow:
  1. Select the TXT file  (multipart POST to /api/nbids-reformat)
  2. Click Generate       (server parses + builds Excel)
  3. Click Download/Save  (decode base64 xlsx → save to Design/New Crew BIds/)
  4. Compare with baseline crew_bids_reference.xlsx  (cell + formatting)

Run: python3 scripts/test_nbids_browser.py
"""
import sys, os, re, json, base64, io, datetime, time, urllib.request, urllib.error

BASE         = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
NBIDS_DIR    = os.path.join(BASE, 'Design', 'New Crew BIds')   # all files live here
SERVER       = 'http://localhost:8088'

# Single-file test cases: (txt filename, expected_period, baseline_xlsx or None)
TEST_CASES = [
    ('Dec Crew Bids 2025 All in one.txt', 'Dec 2025', 'crew_bids_reference.xlsx'),
    ('May Crew Bids 2025 All in one.txt', 'May 2025', None),
]

# Multi-file test case: list of txt filenames uploaded together
MULTI_FILE_CASE = [
    'May Crew Bids 2025 All in one.txt',
    'Dec Crew Bids 2025 All in one.txt',
]


# ── Step 1 & 2: Simulate file select + click Generate ────────────────────────

def post_multipart(url, fields, files):
    """Minimal multipart/form-data POST using stdlib only."""
    boundary = '----PythonBoundary' + str(int(time.time()))
    body_parts = []
    for name, value in fields.items():
        body_parts.append(
            f'--{boundary}\r\n'
            f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
            f'{value}\r\n'.encode()
        )
    for name, (filename, data) in files.items():
        body_parts.append(
            f'--{boundary}\r\n'
            f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'
            f'Content-Type: application/octet-stream\r\n\r\n'.encode()
            + data + b'\r\n'
        )
    body_parts.append(f'--{boundary}--\r\n'.encode())
    body = b''.join(body_parts)
    req = urllib.request.Request(
        url, data=body,
        headers={'Content-Type': f'multipart/form-data; boundary={boundary}',
                 'Content-Length': str(len(body))},
        method='POST'
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode())


def step1_2_upload_and_generate(txt_filenames):
    """Accept one filename string or a list of filenames."""
    if isinstance(txt_filenames, str):
        txt_filenames = [txt_filenames]

    print(f'Step 1: Selecting {len(txt_filenames)} TXT file(s)...')
    files = {}
    for i, fname in enumerate(txt_filenames):
        path = os.path.join(NBIDS_DIR, fname)
        with open(path, 'rb') as f:
            data_bytes = f.read()
        files[f'file_{i}'] = (fname, data_bytes)
        print(f'        {fname}  ({len(data_bytes)//1024} KB)')

    print(f'Step 2: Clicking Generate (POST /api/nbids-reformat, no period field)...')
    # Rebuild as multiple 'file' parts (same field name)
    boundary = '----PythonBoundary' + str(int(time.time()))
    body_parts = []
    body_parts.append(
        f'--{boundary}\r\nContent-Disposition: form-data; name="period"\r\n\r\n\r\n'.encode()
    )
    for fname, data_bytes in files.values():
        body_parts.append(
            f'--{boundary}\r\nContent-Disposition: form-data; name="file"; filename="{fname}"\r\n'
            f'Content-Type: application/octet-stream\r\n\r\n'.encode()
            + data_bytes + b'\r\n'
        )
    body_parts.append(f'--{boundary}--\r\n'.encode())
    body = b''.join(body_parts)

    req = urllib.request.Request(
        f'{SERVER}/api/nbids-reformat', data=body,
        headers={'Content-Type': f'multipart/form-data; boundary={boundary}',
                 'Content-Length': str(len(body))},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode())
    except urllib.error.URLError as e:
        print(f'  ERROR: Cannot reach server at {SERVER} — {e}')
        sys.exit(1)

    print(f'        Files parsed : {result.get("files", len(txt_filenames))}')
    print(f'        Parsed rows  : {result["parsed_rows"]}')
    print(f'        Error rows   : {result["error_rows"]}')
    print(f'        Suggested name: {result["filename"]}')
    return result


# ── Step 3: Simulate Download / Save ─────────────────────────────────────────

def step3_save(data):
    print('Step 3: Downloading and saving Excel...')
    xlsx_bytes = base64.b64decode(data['xlsx_b64'])
    ts       = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
    filename = f'crew_bids_reference-{ts}.xlsx'
    out_path = os.path.join(NBIDS_DIR, filename)
    with open(out_path, 'wb') as f:
        f.write(xlsx_bytes)
    print(f'        Saved: {out_path}')
    return out_path, xlsx_bytes


# ── Step 4: Compare with baseline ────────────────────────────────────────────

def cell_str(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def step4_compare(xlsx_bytes, baseline_filename='crew_bids_reference.xlsx'):
    import openpyxl
    ref_path = os.path.join(NBIDS_DIR, baseline_filename)
    print(f'Step 4: Comparing with baseline ({baseline_filename})...')

    wb_gen = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    wb_ref = openpyxl.load_workbook(ref_path, data_only=True)

    # ── Value comparison ─────────────────────────────────────────────────────
    if 'crew_bids' not in wb_gen.sheetnames:
        print(f"  ERROR: generated file has no 'crew_bids' sheet. Sheets: {wb_gen.sheetnames}")
        return False

    gen_rows = list(wb_gen['crew_bids'].iter_rows(values_only=True))
    ref_rows = list(wb_ref['crew_bids'].iter_rows(values_only=True))

    print(f'  Reference : {len(ref_rows)} rows x {len(ref_rows[0])} cols')
    print(f'  Generated : {len(gen_rows)} rows x {len(gen_rows[0])} cols')

    value_mismatches = []
    for ri in range(max(len(ref_rows), len(gen_rows))):
        if ri >= len(ref_rows):
            value_mismatches.append((ri+1, None, None, 'EXTRA ROW in generated'))
            continue
        if ri >= len(gen_rows):
            value_mismatches.append((ri+1, None, None, 'MISSING ROW in generated'))
            continue
        for ci in range(max(len(ref_rows[ri]), len(gen_rows[ri]))):
            rv = cell_str(ref_rows[ri][ci] if ci < len(ref_rows[ri]) else None)
            gv = cell_str(gen_rows[ri][ci] if ci < len(gen_rows[ri]) else None)
            if rv != gv:
                col = chr(ord('A') + ci) if ci < 26 else f'C{ci}'
                value_mismatches.append((ri+1, col, rv, gv))

    print(f'  Value mismatches: {len(value_mismatches)}')
    if value_mismatches:
        print('\n── VALUE MISMATCHES (first 40) ──')
        for row_n, col, rv, gv in value_mismatches[:40]:
            if col is None:
                print(f'  Row {row_n}: {rv or gv}')
            else:
                print(f'  [{col}{row_n}]  ref={repr(rv)}  gen={repr(gv)}')
        return False

    # ── Formatting comparison ────────────────────────────────────────────────
    wb_gen2 = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    wb_ref2 = openpyxl.load_workbook(ref_path)
    ws_gen2 = wb_gen2['crew_bids']
    ws_ref2 = wb_ref2['crew_bids']

    fmt_mismatches = []

    def _fmt(cell):
        f  = cell.font
        fi = cell.fill
        al = cell.alignment
        fc = (f.color.rgb if f.color and f.color.type == 'rgb' else None) if f.color else None
        fg = (fi.fgColor.rgb if fi.fgColor and fi.fgColor.type == 'rgb' else None) if fi and fi.fgColor else None
        return {'bold': f.bold, 'color': fc, 'size': f.size, 'name': f.name,
                'fill': fi.fill_type if fi else None, 'fgColor': fg,
                'h_align': al.horizontal, 'v_align': al.vertical, 'wrap': al.wrap_text}

    for ri in [1, 2, 3]:
        ref_row = list(ws_ref2.iter_rows(min_row=ri, max_row=ri))[0]
        gen_row = list(ws_gen2.iter_rows(min_row=ri, max_row=ri))[0]
        for rc, gc in zip(ref_row, gen_row):
            rf, gf = _fmt(rc), _fmt(gc)
            for attr in rf:
                if rf[attr] != gf[attr]:
                    fmt_mismatches.append(
                        f'  [{rc.column_letter}{ri}] {attr}: ref={repr(rf[attr])} gen={repr(gf[attr])}'
                    )

    for col in 'ABCDEFGHIJKLMNOPQ':
        rw = ws_ref2.column_dimensions[col].width if col in ws_ref2.column_dimensions else None
        gw = ws_gen2.column_dimensions[col].width if col in ws_gen2.column_dimensions else None
        if rw != gw:
            fmt_mismatches.append(f'  Col {col} width: ref={rw} gen={gw}')

    if ws_ref2.freeze_panes != ws_gen2.freeze_panes:
        fmt_mismatches.append(f'  freeze_panes: ref={ws_ref2.freeze_panes} gen={ws_gen2.freeze_panes}')

    print(f'  Formatting mismatches: {len(fmt_mismatches)}')
    if fmt_mismatches:
        print('\n── FORMATTING MISMATCHES ──')
        for m in fmt_mismatches:
            print(m)
        return False

    return True


# ── Main ─────────────────────────────────────────────────────────────────────

def run_case(txt_filename, expected_period, baseline_xlsx):
    """Run one full browser-simulation test case. Returns True on pass."""
    print(f'\n{"─"*60}')
    print(f'  FILE: {txt_filename}')
    print(f'{"─"*60}')

    data                 = step1_2_upload_and_generate(txt_filename)
    out_path, xlsx_bytes = step3_save(data)

    ok = True

    # ── Period check ──────────────────────────────────────────────────────
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb['crew_bids']
    # period is in column D (index 3), first data row = row 3
    detected_period = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0][3]
    if detected_period == expected_period:
        print(f'  Period  : OK  "{detected_period}"')
    else:
        print(f'  Period  : FAIL  expected="{expected_period}"  got="{detected_period}"')
        ok = False

    # ── Filename check ────────────────────────────────────────────────────
    fname = data['filename']
    if re.match(r'^crew_bids_reference-\d{4}-\d{2}-\d{2}-\d{6}\.xlsx$', fname):
        print(f'  Filename: OK  {fname}')
    else:
        print(f'  Filename: FAIL  "{fname}" does not match standard')
        ok = False

    # ── Baseline comparison (if provided) ────────────────────────────────
    if baseline_xlsx:
        cmp_ok = step4_compare(xlsx_bytes, baseline_xlsx)
        ok = ok and cmp_ok
    else:
        print('  Baseline: skipped (no baseline provided)')

    return ok


def run_multi_file_case(txt_filenames, baseline_filter_period, baseline_xlsx):
    """Upload multiple files together; verify Dec rows match baseline 100%."""
    print(f'\n{"─"*60}')
    print(f'  MULTI-FILE: {", ".join(txt_filenames)}')
    print(f'{"─"*60}')

    import openpyxl

    data                 = step1_2_upload_and_generate(txt_filenames)
    out_path, xlsx_bytes = step3_save(data)
    ok = True

    # Filename check
    fname = data['filename']
    if re.match(r'^crew_bids_reference-\d{4}-\d{2}-\d{2}-\d{6}\.xlsx$', fname):
        print(f'  Filename: OK  {fname}')
    else:
        print(f'  Filename: FAIL  "{fname}"')
        ok = False

    # Verify earlier month appears before later month in the Excel
    wb_gen = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws     = wb_gen['crew_bids']
    periods_in_order = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        p = row[3]  # period column D
        if p and p not in periods_in_order:
            periods_in_order.append(p)
    print(f'  Period order in Excel: {periods_in_order}')
    if len(periods_in_order) >= 2:
        _MO = {m:i for i,m in enumerate(['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'])}
        def _mo(p):
            parts = p.split(); return (_MO.get(parts[0][:3].capitalize(),99), int(parts[1]) if len(parts)>1 else 9999)
        if _mo(periods_in_order[0]) <= _mo(periods_in_order[1]):
            print(f'  Sort order: OK  (earlier month first)')
        else:
            print(f'  Sort order: FAIL  (later month appears first)')
            ok = False

    # Compare Dec rows in combined Excel against Dec baseline
    print(f'\nStep 4: Comparing {baseline_filter_period} rows against baseline ({baseline_xlsx})...')
    wb_ref = openpyxl.load_workbook(os.path.join(NBIDS_DIR, baseline_xlsx), data_only=True)
    ref_rows = list(wb_ref['crew_bids'].iter_rows(values_only=True))  # incl header + desc

    # Extract only rows matching the filter period from generated Excel
    gen_all  = list(ws.iter_rows(min_row=1, values_only=True))
    gen_header = gen_all[0]
    gen_desc   = gen_all[1]
    gen_data   = [r for r in gen_all[2:] if r[3] == baseline_filter_period]  # col D = period
    gen_subset = [gen_header, gen_desc] + gen_data

    print(f'  Reference rows (crew_bids): {len(ref_rows)}')
    print(f'  Generated {baseline_filter_period} rows: {len(gen_subset)}')

    # Column F (index 5) = property_group_id — expected to differ in multi-file
    # because it is offset by the preceding file's group count; skip it
    SKIP_COLS = {5}

    mismatches = []
    for ri in range(max(len(ref_rows), len(gen_subset))):
        if ri >= len(ref_rows):
            mismatches.append((ri+1, None, None, 'EXTRA ROW in generated subset')); continue
        if ri >= len(gen_subset):
            mismatches.append((ri+1, None, None, 'MISSING ROW in generated subset')); continue
        for ci in range(max(len(ref_rows[ri]), len(gen_subset[ri]))):
            if ci in SKIP_COLS:
                continue
            rv = cell_str(ref_rows[ri][ci] if ci < len(ref_rows[ri]) else None)
            gv = cell_str(gen_subset[ri][ci] if ci < len(gen_subset[ri]) else None)
            if rv != gv:
                col = chr(ord('A')+ci) if ci < 26 else f'C{ci}'
                mismatches.append((ri+1, col, rv, gv))

    print(f'  Value mismatches: {len(mismatches)}')
    if mismatches:
        print(f'\n── MISMATCHES (first 20) ──')
        for row_n, col, rv, gv in mismatches[:20]:
            if col is None: print(f'  Row {row_n}: {rv or gv}')
            else:           print(f'  [{col}{row_n}]  ref={repr(rv)}  gen={repr(gv)}')
        ok = False
    else:
        print(f'  {baseline_filter_period} rows: PERFECT MATCH -- 100%')

    return ok


def main():
    print('=' * 60)
    print('  N-Bids Browser Simulation Test')
    print('=' * 60)

    results = {}

    # ── Round 1: Upload May + Dec together, compare Dec with baseline ─────────
    print(f'\n{"#"*60}')
    print('  ROUND 1 — Upload May + Dec, compare Dec with baseline')
    print(f'{"#"*60}')
    results['round1'] = run_multi_file_case(
        ['May Crew Bids 2025 All in one.txt', 'Dec Crew Bids 2025 All in one.txt'],
        baseline_filter_period='Dec 2025',
        baseline_xlsx='crew_bids_reference.xlsx'
    )

    # ── Round 2: Upload Dec only, compare Dec with baseline ───────────────────
    print(f'\n{"#"*60}')
    print('  ROUND 2 — Upload Dec only, compare Dec with baseline')
    print(f'{"#"*60}')
    results['round2'] = run_multi_file_case(
        ['Dec Crew Bids 2025 All in one.txt'],
        baseline_filter_period='Dec 2025',
        baseline_xlsx='crew_bids_reference.xlsx'
    )

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f'\n{"="*60}')
    print(f'  ROUND 1 (May+Dec → Dec vs baseline): {"PASS ✓" if results["round1"] else "FAIL ✗"}')
    print(f'  ROUND 2 (Dec only → Dec vs baseline): {"PASS ✓" if results["round2"] else "FAIL ✗"}')
    all_ok = all(results.values())
    print(f'\n  {"ALL ROUNDS PASSED -- 100%" if all_ok else "SOME ROUNDS FAILED"}')
    sys.exit(0 if all_ok else 1)


if __name__ == '__main__':
    main()
